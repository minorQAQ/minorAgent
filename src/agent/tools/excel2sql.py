"""excel2sql 工具：按大模型给定的数据区范围，自动建库建表并把 Excel 数据批量入库。

系统定位:
    配合 hard_excel_read 完成"解析 → 入库"链路。hard_excel_read 让大模型理解复杂表结构、
    产出 data_region 与 table_description；本工具按该范围流式读取真正的数据区，
    自动建库建表（表 COMMENT 存表格描述，仅 MySQL；SQLite 无 COMMENT 概念）并批量入库。

    连接使用设置中的**工具数据库配置**（TOOL_DB_*，MySQL 或 SQLite），
    与存储后端配置完全隔离。SQLite 目录模式下库名 = 文件名（<目录>/<库名>.sqlite）。

    不做重清洗 DSL，按范围直接入库（结构清洗已由大模型定范围完成）。
    值级清洗入库后用 text2sql 的 execute(UPDATE/DELETE) 完成。

权限策略（见 tool_config.json）:
    permission=confirm（写库）；dry_run="true" 时经 auto_execute_rules 升级 direct。
    注意：auto_execute_rules 的 equals 是 str(actual)==str(expected)，
    故 dry_run 用字符串字段，配 value="true"（bool 的 str(True)=="True" 会匹配失败）。
"""
from __future__ import annotations

import datetime
import json
import os
import re
from contextlib import contextmanager
from typing import Any

from langchain.tools import BaseTool
from pydantic import BaseModel, Field


@contextmanager
def _cursor(conn: Any):
    """统一游标上下文：pymysql 与 sqlite3 的 cursor 行为差异。

    pymysql cursor 支持 with（自动 close），sqlite3 cursor 不支持
    context manager 协议，需手动 close。本助手对两者统一处理。
    """
    cur = conn.cursor()
    try:
        yield cur
    finally:
        try:
            cur.close()
        except Exception:
            pass


# data_region 必须包含的键
_REGION_KEYS = ("header_row", "data_start", "data_end", "col_start", "col_end")


def _file_is_xls(file_path: str) -> bool:
    return file_path.lower().endswith(".xls") and not file_path.lower().endswith(".xlsx")


def _xlrd_cell_to_value(sheet, r: int, c: int) -> Any:
    """xlrd 单元格值转 Python 原生类型（处理日期序列号）。"""
    import xlrd
    ctype = sheet.cell_type(r, c)
    val = sheet.cell_value(r, c)
    if ctype == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate_as_datetime(val, sheet.book.datemode)
        except Exception:
            return val
    if ctype == xlrd.XL_CELL_EMPTY:
        return None
    if ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(val)
    return val


# ---------- 库名/表名校验（与 text2sql 一致） ----------
_DB_NAME_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")
_TABLE_NAME_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")


def _db_type() -> str:
    """当前工具数据库类型："mysql" | "sqlite"（读 TOOL_DB_*，与存储后端隔离）。"""
    from agent.core.db import load_tool_db_config
    return load_tool_db_config().get("type", "mysql")


def _ensure_database_exists(db_name: str) -> tuple[bool, str]:
    """幂等创建目标数据库。

    - mysql: 连服务器（不指定 database）执行 CREATE DATABASE IF NOT EXISTS；
    - sqlite: 库 = 文件，幂等创建（复用 ensure_sqlite_database，不存在则建空文件）。
    """
    if _db_type() == "sqlite":
        from agent.core.db import load_tool_db_config, sqlite_db_path, ensure_sqlite_database
        try:
            cfg = load_tool_db_config()
            path = sqlite_db_path(cfg, db_name)
            ensure_sqlite_database(path)
            return True, "ok"
        except Exception as e:
            return False, str(e)
    import pymysql
    from pymysql.cursors import DictCursor
    from agent.core.db import load_tool_db_config

    cfg = load_tool_db_config()
    try:
        conn = pymysql.connect(
            host=cfg["host"], port=cfg["port"], user=cfg["user"],
            password=cfg["password"], charset="utf8mb4",
            cursorclass=DictCursor, autocommit=True,
        )
        try:
            with _cursor(conn) as cur:
                cur.execute(
                    f"CREATE DATABASE IF NOT EXISTS `{db_name}` "
                    f"CHARACTER SET utf8mb4 COLLATE utf8mb4_general_ci"
                )
        finally:
            conn.close()
        return True, "ok"
    except Exception as e:
        return False, str(e)


def _connect(db_name: str):
    """连接目标库（按工具数据库配置分发 mysql/sqlite）。"""
    from agent.core.db import connect_db, load_tool_db_config

    cfg = load_tool_db_config()
    return connect_db(cfg, db_name)


def _table_exists(conn: Any, table: str) -> bool:
    """判断目标表是否已存在（mysql: SHOW TABLES LIKE；sqlite: sqlite_master）。"""
    if _db_type() == "sqlite":
        cur = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table,)
        )
        return cur.fetchone() is not None
    with _cursor(conn) as cur:
        cur.execute(f"SHOW TABLES LIKE '{table}'")
        return cur.fetchone() is not None


def _placeholder() -> str:
    """SQL 占位符：mysql 用 %s，sqlite 用 ?。"""
    return "?" if _db_type() == "sqlite" else "%s"


def _normalize_value(v: Any) -> Any:
    """单元格值规范化：空字符串转 None；其余原样（pymysql 能处理 int/float/datetime/str）。"""
    if v is None:
        return None
    if isinstance(v, str) and v.strip() == "":
        return None
    if isinstance(v, str):
        return v.strip()
    return v


def _is_empty(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def _infer_sql_type(values: list) -> str:
    """根据列样本值推断 MySQL 类型。"""
    non_empty = [v for v in values if not _is_empty(v)]
    if not non_empty:
        return "TEXT"
    # 全 int（非 bool）
    if all(isinstance(v, int) and not isinstance(v, bool) for v in non_empty):
        return "BIGINT"
    # 全 numeric
    if all(isinstance(v, (int, float)) and not isinstance(v, bool) for v in non_empty):
        return "DECIMAL(18,4)"
    # 全 datetime
    if all(isinstance(v, datetime.datetime) for v in non_empty):
        return "DATETIME"
    if all(isinstance(v, datetime.date) for v in non_empty):
        return "DATE"
    # 文本
    max_len = max(len(str(v)) for v in non_empty)
    if max_len > 200:
        return "TEXT"
    n = max(16, ((int(max_len * 1.5) + 15) // 16) * 16)
    n = min(n, 500)
    return f"VARCHAR({n})"


def _clean_col_name(raw: str, idx: int, used: set) -> str:
    """清洗列名为可用标识符（保留中文，去控制字符；空/重复则兜底）。"""
    if raw is None:
        name = ""
    else:
        name = str(raw).strip()
        # 去控制字符（换行/制表符等）
        name = re.sub(r"[\x00-\x1f\x7f]", "", name)
    if not name:
        name = f"col_{idx + 1}"
    # 去重
    base = name
    k = 2
    while name in used:
        name = f"{base}_{k}"
        k += 1
    used.add(name)
    return name


def _resolve_columns(columns: list[dict], header_values: list, used: set) -> list[dict]:
    """按用户给的 columns 定义，映射到数据区列位置。

    columns 每项: {source, target, sql_type, nullable?, primary_key?, index?, comment?}
    source 必须匹配 header_values 中的某个值，否则报错。
    """
    # 建 source -> pos 映射（按值匹配，取首个）
    header_map: dict = {}
    for i, v in enumerate(header_values):
        key = "" if v is None else str(v).strip()
        if key and key not in header_map:
            header_map[key] = i

    col_defs: list[dict] = []
    for c in columns:
        src = c.get("source")
        src_key = "" if src is None else str(src).strip()
        if src_key not in header_map:
            raise ValueError(
                f"columns.source '{src}' 在表头中未找到。可用表头: {header_values}"
            )
        pos = header_map[src_key]
        target = _clean_col_name(c.get("target") or src_key, pos, used)
        col_defs.append({
            "pos": pos,
            "source_name": src_key,
            "target": target,
            "sql_type": c.get("sql_type") or "VARCHAR(255)",
            "nullable": c.get("nullable", True),
            "primary_key": c.get("primary_key", False),
            "index": c.get("index", False),
            "comment": c.get("comment", ""),
        })
    return col_defs


def _infer_columns(header_values: list, sample_rows: list) -> list[dict]:
    """无 columns 定义时，用表头作列名 + 抽样推断类型。"""
    used: set = set()
    col_defs: list[dict] = []
    for i, raw_name in enumerate(header_values):
        target = _clean_col_name(raw_name, i, used)
        col_vals = [r[i] if i < len(r) else None for r in sample_rows]
        sql_type = _infer_sql_type(col_vals)
        col_defs.append({
            "pos": i,
            "source_name": "" if raw_name is None else str(raw_name).strip(),
            "target": target,
            "sql_type": sql_type,
            "nullable": True,
            "primary_key": False,
            "index": False,
            "comment": "",
        })
    return col_defs


def _build_ddl(table: str, col_defs: list[dict], description: str,
               drop_if_exists: bool, db_type: str | None = None) -> str:
    """生成建表 DDL（按数据库类型区分语法）。

    - mysql: 列 COMMENT / 表 COMMENT / ENGINE=InnoDB DEFAULT CHARSET / 行内 KEY 索引；
    - sqlite: 无 COMMENT / ENGINE / 行内 KEY（索引改为建表后独立
      CREATE INDEX IF NOT EXISTS 语句）。
    列类型（BIGINT/DECIMAL/DATETIME/DATE/TEXT/VARCHAR）两者通用（sqlite 亲和类型）。
    """
    db_type = db_type or _db_type()
    is_sqlite = db_type == "sqlite"
    lines: list[str] = []
    if drop_if_exists:
        lines.append(f"DROP TABLE IF EXISTS `{table}`;")

    col_lines: list[str] = []
    pks: list[str] = []
    for c in col_defs:
        parts = [f"  `{c['target']}` {c['sql_type']}"]
        if not c.get("nullable", True):
            parts.append("NOT NULL")
        if c.get("comment") and not is_sqlite:
            esc = str(c["comment"]).replace("'", "''")[:500]
            parts.append(f"COMMENT '{esc}'")
        col_lines.append(" ".join(parts))
        if c.get("primary_key"):
            pks.append(c["target"])
    if pks:
        col_lines.append(
            f"  PRIMARY KEY ({', '.join(f'`{p}`' for p in pks)})"
        )
    if not is_sqlite:
        for c in col_defs:
            if c.get("index") and not c.get("primary_key"):
                idx_name = f"idx_{c['target']}"[:64]
                col_lines.append(f"  KEY `{idx_name}` (`{c['target']}`)")

    if is_sqlite:
        ddl = (
            f"CREATE TABLE IF NOT EXISTS `{table}` (\n"
            + ",\n".join(col_lines)
            + "\n)"
        )
        # sqlite 索引在表外单独建（IF NOT EXISTS 幂等）
        for c in col_defs:
            if c.get("index") and not c.get("primary_key"):
                idx_name = f"idx_{c['target']}"[:64]
                lines.append(
                    f"CREATE INDEX IF NOT EXISTS `{idx_name}` ON `{table}` (`{c['target']}`);"
                )
    else:
        desc_clause = ""
        if description:
            esc = description.replace("'", "''")[:1000]
            desc_clause = f" COMMENT='{esc}'"
        ddl = (
            f"CREATE TABLE IF NOT EXISTS `{table}` (\n"
            + ",\n".join(col_lines)
            + f"\n) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4{desc_clause}"
        )
    lines.append(ddl)
    return "\n".join(lines)


def _row_to_dict(row: tuple, col_defs: list[dict]) -> dict:
    return {c["target"]: _normalize_value(row[c["pos"]] if c["pos"] < len(row) else None)
            for c in col_defs}


def _row_to_tuple(row: tuple, col_defs: list[dict]) -> list:
    return [_normalize_value(row[c["pos"]] if c["pos"] < len(row) else None)
            for c in col_defs]


class Excel2SQLInput(BaseModel):
    file_path: str = Field(..., description="Excel 文件绝对路径")
    sheet: str = Field(..., description="要入库的 sheet 名")
    data_region: dict = Field(
        ...,
        description="行列式数据区范围 {header_row, data_start, data_end, col_start, col_end}（1-indexed），"
                    "由大模型从 hard_excel_read 给出",
    )
    table_description: str = Field(
        "",
        description="表格描述（数据区上方的标题/表单/说明等），存为表 COMMENT",
    )
    target_database: str = Field(..., description="目标库名（不存在自动建；连接使用工具数据库配置，MySQL 或 SQLite）")
    target_table: str = Field(..., description="目标表名（不存在自动建）")
    columns: list[dict] | None = Field(
        None,
        description="列定义 [{source,target,sql_type,nullable?,primary_key?,index?,comment?}]；"
                    "None=用数据区表头作列名+自动推断类型",
    )
    batch_size: int = Field(2000, description="executemany 批量大小")
    drop_if_exists: bool = Field(
        False,
        description="是否先 DROP TABLE（重建）；默认 False 走 CREATE IF NOT EXISTS + INSERT",
    )
    dry_run: str = Field(
        "false",
        description="'true'=只读不写库，返回样本与 DDL 预览供复核；'false'=入库",
    )
    sample_after: int = Field(5, description="返回清洗后前 N 行样本")


class Excel2SQLTool(BaseTool):
    """按数据区范围自动建库建表 + 批量入库 Excel 数据。"""

    args_schema: type[BaseModel] = Excel2SQLInput
    name: str = "excel2sql"
    description: str = (
        "Excel 批量入库工具（excel2sql）。按大模型给定的数据行列范围，自动建库建表"
        "（表描述存为表 COMMENT，仅 MySQL；SQLite 无此概念），"
        "并把 Excel 数据区流式批量入库（MySQL 或 SQLite，连接使用工具数据库配置，与存储库隔离）。\n"
        "适用场景：hard_excel_read 已解析出 data_region 与 table_description 后，把数据真正写入数据库。\n"
        "不做重清洗 DSL，按范围直接入库；值级清洗入库后用 text2sql(execute) 完成。\n"
        "参数说明：\n"
        "- file_path: Excel 文件绝对路径\n"
        "- sheet: 要入库的 sheet 名\n"
        "- data_region: {header_row, data_start, data_end, col_start, col_end}（1-indexed）\n"
        "- table_description: 表格描述，存为表 COMMENT\n"
        "- target_database: 目标库名（不存在自动建）\n"
        "- target_table: 目标表名（不存在自动建）\n"
        "- columns: 列定义 [{source,target,sql_type,nullable,primary_key,index,comment}]；None=表头+自动推断\n"
        "- batch_size: 批量入库大小，默认 2000\n"
        "- drop_if_exists: 是否先 DROP 重建，默认 False\n"
        "- dry_run: 'true'=只读复核（direct）；'false'=入库（confirm）\n"
        "- sample_after: 返回前 N 行样本\n"
        "权限：dry_run='true' 直接执行；dry_run='false' 需用户确认。\n"
        "调用示例：\n"
        '- 复核: {"file_path":"C:\\\\d.xlsx","sheet":"S1","data_region":{"header_row":4,"data_start":5,"data_end":8504,"col_start":1,"col_end":28},"target_database":"sales","target_table":"sales_2023","dry_run":"true"}\n'
    )

    def _run(
        self,
        file_path: str,
        sheet: str,
        data_region: dict,
        table_description: str = "",
        target_database: str | None = None,
        target_table: str | None = None,
        columns: list[dict] | None = None,
        batch_size: int = 2000,
        drop_if_exists: bool = False,
        dry_run: str = "false",
        sample_after: int = 5,
        **kwargs: Any,
    ) -> str:
        try:
            # ---- 基础校验 ----
            if not os.path.isfile(file_path):
                return json.dumps({"error": f"文件不存在: {file_path}"}, ensure_ascii=False)
            if not target_database or not _DB_NAME_RE.match(target_database):
                return json.dumps({"error": f"非法目标库名: {target_database}"}, ensure_ascii=False)
            if not target_table or not _TABLE_NAME_RE.match(target_table):
                return json.dumps({"error": f"非法目标表名: {target_table}"}, ensure_ascii=False)
            if not isinstance(data_region, dict) or not all(k in data_region for k in _REGION_KEYS):
                return json.dumps(
                    {"error": f"data_region 必须包含 {_REGION_KEYS}"}, ensure_ascii=False,
                )

            header_row = int(data_region["header_row"])
            data_start = int(data_region["data_start"])
            data_end = int(data_region["data_end"])
            col_start = int(data_region["col_start"])
            col_end = int(data_region["col_end"])
            if data_end < data_start or col_end < col_start:
                return json.dumps({"error": "data_region 范围非法: data_end<data_start 或 col_end<col_start"}, ensure_ascii=False)

            # 按文件格式选择引擎
            if _file_is_xls(file_path):
                return self._run_xls(
                    file_path, sheet, data_region, table_description,
                    target_database, target_table, columns,
                    batch_size, drop_if_exists, dry_run, sample_after,
                )

            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            try:
                if sheet not in wb.sheetnames:
                    return json.dumps(
                        {"error": f"sheet 不存在: {sheet}", "available": wb.sheetnames},
                        ensure_ascii=False,
                    )
                ws = wb[sheet]

                # ---- 读表头 ----
                header_values: list = []
                for row in ws.iter_rows(min_row=header_row, max_row=header_row,
                                        min_col=col_start, max_col=col_end, values_only=True):
                    header_values = list(row)
                    break
                if not header_values:
                    return json.dumps({"error": f"表头行 {header_row} 为空"}, ensure_ascii=False)

                # ---- 确定列定义 ----
                used: set = set()
                try:
                    if columns:
                        col_defs = _resolve_columns(columns, header_values, used)
                    else:
                        # 抽样前 200 行用于类型推断
                        sample_n = min(200, data_end - data_start + 1)
                        sample_rows: list = []
                        for row in ws.iter_rows(min_row=data_start, max_row=data_start + sample_n - 1,
                                                min_col=col_start, max_col=col_end, values_only=True):
                            sample_rows.append(list(row))
                        col_defs = _infer_columns(header_values, sample_rows)
                except ValueError as e:
                    return json.dumps({"error": str(e)}, ensure_ascii=False)

                # ---- 生成 DDL ----
                ddl = _build_ddl(target_table, col_defs, table_description, drop_if_exists)

                # ---- 检测表是否存在 + schema 冲突（仅 dry_run=false 且 columns 明确指定时）----
                if dry_run != "true" and columns and not drop_if_exists:
                    try:
                        ok, _ = _ensure_database_exists(target_database)
                        if ok:
                            conn = _connect(target_database)
                            try:
                                if _table_exists(conn, target_table):
                                    return json.dumps({
                                            "error": (
                                                f"表 `{target_database}`.`{target_table}` 已存在，"
                                                f"而当前传入的 columns 定义了新的列 schema，"
                                                f"CREATE TABLE IF NOT EXISTS 不会修改已有表结构。"
                                                f"请设置 drop_if_exists: true 以删表重建，"
                                                f"或先 dry_run=true 复核再决定。"
                                            ),
                                            "hint": "设置 drop_if_exists: true",
                                        }, ensure_ascii=False)
                            finally:
                                conn.close()
                    except Exception:
                        pass  # 连不上库时不阻断，让后续入库失败时自然报错

                # ---- dry_run: 只返回样本 + DDL 预览 ----
                if dry_run == "true":
                    sample_dicts: list[dict] = []
                    for row in ws.iter_rows(min_row=data_start,
                                            max_row=min(data_start + sample_after - 1, data_end),
                                            min_col=col_start, max_col=col_end, values_only=True):
                        sample_dicts.append(_row_to_dict(row, col_defs))
                        if len(sample_dicts) >= sample_after:
                            break
                    return json.dumps({
                        "status": "dry_run",
                        "database": target_database,
                        "table": target_table,
                        "data_region": data_region,
                        "columns": [{"target": c["target"], "sql_type": c["sql_type"],
                                     "source": c["source_name"]} for c in col_defs],
                        "ddl_preview": ddl,
                        "sample_after": sample_dicts,
                    }, ensure_ascii=False, default=str)

                # ---- 正式入库 ----
                # 1) 建库
                ok, msg = _ensure_database_exists(target_database)
                if not ok:
                    return json.dumps(
                        {"error": f"建库失败: {msg}", "database": target_database},
                        ensure_ascii=False,
                    )
                # 2) 建表 + 入库
                conn = _connect(target_database)
                inserted = 0
                rejected = 0
                batches = 0
                sample_dicts: list[dict] = []
                batch: list = []
                try:
                    with _cursor(conn) as cur:
                        # 执行 DDL（可能含 DROP + CREATE）
                        for stmt in ddl.split(";"):
                            s = stmt.strip()
                            if s:
                                cur.execute(s)
                    # 入库
                    cols_sql = ", ".join(f"`{c['target']}`" for c in col_defs)
                    placeholders = ", ".join([_placeholder()] * len(col_defs))
                    insert_sql = f"INSERT INTO `{target_table}` ({cols_sql}) VALUES ({placeholders})"

                    with _cursor(conn) as cur:
                        for row in ws.iter_rows(min_row=data_start, max_row=data_end,
                                                min_col=col_start, max_col=col_end, values_only=True):
                            # 整行空跳过
                            if all(_is_empty(v) for v in row):
                                rejected += 1
                                continue
                            batch.append(_row_to_tuple(row, col_defs))
                            if len(sample_dicts) < sample_after:
                                sample_dicts.append(_row_to_dict(row, col_defs))
                            if len(batch) >= batch_size:
                                cur.executemany(insert_sql, batch)
                                inserted += cur.rowcount
                                conn.commit()
                                batches += 1
                                batch = []
                        if batch:
                            cur.executemany(insert_sql, batch)
                            inserted += cur.rowcount
                            conn.commit()
                            batches += 1
                except Exception:
                    try:
                        conn.rollback()
                    except Exception:
                        pass
                    raise
                finally:
                    conn.close()

                return json.dumps({
                    "status": "loaded",
                    "database": target_database,
                    "table": target_table,
                    "inserted": inserted,
                    "rejected": rejected,
                    "batches": batches,
                    "sample_after": sample_dicts,
                    "ddl_used": ddl,
                    "table_description_saved": bool(table_description),
                }, ensure_ascii=False, default=str)
            finally:
                wb.close()
        except Exception as e:
            return json.dumps({"error": f"excel2sql 失败: {str(e)}"}, ensure_ascii=False)

    def _run_xls(
        self,
        file_path: str,
        sheet: str,
        data_region: dict,
        table_description: str,
        target_database: str | None,
        target_table: str | None,
        columns: list[dict] | None,
        batch_size: int,
        drop_if_exists: bool,
        dry_run: str,
        sample_after: int,
    ) -> str:
        """xlrd 引擎：处理 .xls 文件入库。"""
        import xlrd
        book = xlrd.open_workbook(file_path)
        if sheet not in book.sheet_names():
            return json.dumps(
                {"error": f"sheet 不存在: {sheet}", "available": book.sheet_names()},
                ensure_ascii=False,
            )
        ws = book.sheet_by_name(sheet)
        header_row = int(data_region["header_row"])
        data_start = int(data_region["data_start"])
        data_end = int(data_region["data_end"])
        col_start = int(data_region["col_start"])
        col_end = int(data_region["col_end"])
        total_rows = ws.nrows

        # 读表头（xlrd: 行索引从 0 开始）
        hr = header_row - 1
        header_values = [_xlrd_cell_to_value(ws, hr, c) for c in range(col_start - 1, min(col_end, ws.ncols))]
        if not header_values:
            return json.dumps({"error": f"表头行 {header_row} 为空"}, ensure_ascii=False)

        # 确定列定义
        used: set = set()
        try:
            if columns:
                col_defs = _resolve_columns(columns, header_values, used)
            else:
                sample_n = min(200, data_end - data_start + 1)
                sample_rows: list = []
                for r in range(data_start - 1, min(data_start - 1 + sample_n, total_rows)):
                    row = [_xlrd_cell_to_value(ws, r, c)
                           for c in range(col_start - 1, min(col_end, ws.ncols))]
                    sample_rows.append(row)
                col_defs = _infer_columns(header_values, sample_rows)
        except ValueError as e:
            return json.dumps({"error": str(e)}, ensure_ascii=False)

        ddl = _build_ddl(target_table, col_defs, table_description, drop_if_exists)

        # 检测表已存在且 columns 明确指定时强制要求 drop_if_exists
        if dry_run != "true" and columns and not drop_if_exists:
            try:
                ok, _ = _ensure_database_exists(target_database)
                if ok:
                    conn = _connect(target_database)
                    try:
                        if _table_exists(conn, target_table):
                            return json.dumps({
                                "error": (
                                    f"表 `{target_database}`.`{target_table}` 已存在，"
                                    f"而当前传入的 columns 定义了新的列 schema，"
                                    f"CREATE TABLE IF NOT EXISTS 不会修改已有表结构。"
                                    f"请设置 drop_if_exists: true 以删表重建。"
                                ),
                                "hint": "设置 drop_if_exists: true",
                            }, ensure_ascii=False)
                    finally:
                        conn.close()
            except Exception:
                pass

        # dry_run
        if dry_run == "true":
            sample_dicts: list[dict] = []
            for r in range(data_start - 1, min(data_start - 1 + sample_after, total_rows)):
                row = tuple(_xlrd_cell_to_value(ws, r, c)
                           for c in range(col_start - 1, min(col_end, ws.ncols)))
                sample_dicts.append(_row_to_dict(row, col_defs))
            return json.dumps({
                "status": "dry_run",
                "database": target_database,
                "table": target_table,
                "data_region": data_region,
                "engine": "xlrd",
                "columns": [{"target": c["target"], "sql_type": c["sql_type"],
                             "source": c["source_name"]} for c in col_defs],
                "ddl_preview": ddl,
                "sample_after": sample_dicts,
            }, ensure_ascii=False, default=str)

        # 正式入库
        ok, msg = _ensure_database_exists(target_database)
        if not ok:
            return json.dumps(
                {"error": f"建库失败: {msg}", "database": target_database},
                ensure_ascii=False,
            )
        conn = _connect(target_database)
        inserted = 0
        rejected = 0
        batches = 0
        sample_dicts: list[dict] = []
        batch: list = []
        try:
            with _cursor(conn) as cur:
                for stmt in ddl.split(";"):
                    s = stmt.strip()
                    if s:
                        cur.execute(s)
            cols_sql = ", ".join(f"`{c['target']}`" for c in col_defs)
            placeholders = ", ".join([_placeholder()] * len(col_defs))
            insert_sql = f"INSERT INTO `{target_table}` ({cols_sql}) VALUES ({placeholders})"

            with _cursor(conn) as cur:
                for r in range(data_start - 1, min(data_end, total_rows)):
                    row_vals = tuple(_xlrd_cell_to_value(ws, r, c)
                                    for c in range(col_start - 1, min(col_end, ws.ncols)))
                    if all(_is_empty(v) for v in row_vals):
                        rejected += 1
                        continue
                    batch.append(_row_to_tuple(row_vals, col_defs))
                    if len(sample_dicts) < sample_after:
                        sample_dicts.append(_row_to_dict(row_vals, col_defs))
                    if len(batch) >= batch_size:
                        cur.executemany(insert_sql, batch)
                        inserted += cur.rowcount
                        conn.commit()
                        batches += 1
                        batch = []
                if batch:
                    cur.executemany(insert_sql, batch)
                    inserted += cur.rowcount
                    conn.commit()
                    batches += 1
        except Exception:
            try:
                conn.rollback()
            except Exception:
                pass
            raise
        finally:
            conn.close()

        return json.dumps({
            "status": "loaded",
            "database": target_database,
            "table": target_table,
            "engine": "xlrd",
            "inserted": inserted,
            "rejected": rejected,
            "batches": batches,
            "sample_after": sample_dicts,
            "ddl_used": ddl,
            "table_description_saved": bool(table_description),
        }, ensure_ascii=False, default=str)

    async def _arun(self, **kwargs: Any) -> str:
        import asyncio
        return await asyncio.to_thread(self._run, **kwargs)
