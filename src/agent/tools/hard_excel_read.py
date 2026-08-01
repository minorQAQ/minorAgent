"""hard_excel_read 工具：读取 Excel sheet 前 N 行原始数据 + 辅助摘要，供大模型解析复杂表结构。

系统定位:
    复杂 Excel 表（含标题区/表单区/行列式数据区/尾部区）的结构理解辅助。
    把 sheet 前 N 行原始数据交给大模型，由大模型（而非启发式算法）理解结构，
    产出"表格描述 + 行列式数据区范围"。工具不内部调 LLM，只返回数据。

    绝不返回数据区全量行，只返回前 N 行 + 辅助摘要，控制 LLM 上下文体积。
    配合 excel2sql 工具完成"解析 → 入库"链路。

权限策略（见 tool_config.json）:
    只读工具，permission=direct。
"""
from __future__ import annotations

import datetime
import json
import os
import zipfile
import xml.etree.ElementTree as _ET
from typing import Any

from langchain.tools import BaseTool
from pydantic import BaseModel, Field


# 单行 preview 文本最大字符数
_MAX_ROW_CHARS = 500
# 单个工具返回 JSON 最大字符数（超限则截断 sheet 列表并提示分批）
_MAX_RETURN_CHARS = 110000


def _col_letter(idx: int) -> str:
    """1-indexed 列号转字母（1->A, 27->AA）。"""
    from openpyxl.utils import get_column_letter
    return get_column_letter(idx)


def _col_letter_from_index(idx: int) -> str:
    """0-indexed 列号转字母（0->A, 26->AA），避免 openpyxl 依赖。"""
    s = ""
    n = idx + 1
    while n > 0:
        n -= 1
        s = chr(ord("A") + n % 26) + s
        n //= 26
    return s


def _file_is_xls(file_path: str) -> bool:
    """判断文件是否为旧版 .xls（通过扩展名）。"""
    return file_path.lower().endswith(".xls") and not file_path.lower().endswith(".xlsx")


def _load_workbook_xls(file_path: str):
    """用 xlrd 加载 .xls 文件，返回 (book, engine_label)。"""
    import xlrd
    return xlrd.open_workbook(file_path), "xlrd"


def _xlrd_cell_to_value(sheet, r: int, c: int):
    """xlrd 单元格值转 Python 原生类型（处理日期序列号等）。"""
    import xlrd
    from datetime import datetime, timedelta
    ctype = sheet.cell_type(r, c)
    val = sheet.cell_value(r, c)
    if ctype == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate_as_datetime(val, sheet.book.datemode)
        except Exception:
            return val
    if ctype == xlrd.XL_CELL_EMPTY:
        return None
    return val


def _get_xls_merged(sheet, preview_rows: int):
    """从 xlrd Sheet 提取合并单元格（仅扫描前 preview_rows 行）。
    返回 (merged_ranges: list[str], merged_top_lefts: set)。

    xlrd merged_cells: [(rlo, rhi, clo, chi)] 均为 0-based，rhi/chi 是 exclusive。
    """
    merged_ranges: list[str] = []
    merged_top_lefts: set = set()
    for rlo, rhi, clo, chi in sheet.merged_cells:
        if rlo < preview_rows:
            ref = f"{_col_letter_from_index(clo)}{rlo + 1}:{_col_letter_from_index(chi - 1)}{rhi}"
            merged_ranges.append(ref)
            merged_top_lefts.add((rlo + 1, clo + 1))
    return merged_ranges, merged_top_lefts


# xlsx 命名空间
_NS = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
_R_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"


def _read_all_merged_cells(file_path: str) -> dict[str, list[str]]:
    """直接从 xlsx(zip) 解析所有 sheet 的合并单元格区域。

    read_only 模式下 openpyxl 不解析 mergeCells 元数据（ws.merged_cells.ranges 为空），
    故用 zipfile 直接读 sheetN.xml 的 <mergeCell ref=.../>。返回 {sheet_name: [ref,...]}。
    失败返回空 dict（不影响主流程）。
    """
    result: dict[str, list[str]] = {}
    try:
        with zipfile.ZipFile(file_path, "r") as z:
            # 1) workbook.xml: sheet name -> r:id，以及 sheet 顺序
            wb_root = _ET.fromstring(z.read("xl/workbook.xml"))
            name_to_rid: dict[str, str] = {}
            for s in wb_root.findall(".//main:sheet", _NS):
                name = s.attrib.get("name", "")
                rid = s.attrib.get(_R_NS, "")
                if name and rid:
                    name_to_rid[name] = rid
            # 2) workbook.xml.rels: r:id -> target(worksheets/sheetN.xml)
            rels_root = _ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
            rid_to_target: dict[str, str] = {}
            for rel in rels_root:
                rid_to_target[rel.attrib.get("Id", "")] = rel.attrib.get("Target", "")
            # 3) 每个 sheet xml: 读 <mergeCell ref=.../>
            for name, rid in name_to_rid.items():
                target = rid_to_target.get(rid, "")
                if not target:
                    continue
                if target.startswith("/"):
                    sheet_path = target[1:]
                else:
                    sheet_path = "xl/" + target
                try:
                    sroot = _ET.fromstring(z.read(sheet_path))
                except KeyError:
                    continue
                refs = [mc.attrib.get("ref", "") for mc in sroot.findall(".//main:mergeCell", _NS)]
                refs = [r for r in refs if r]
                if refs:
                    result[name] = refs
    except Exception:
        pass
    return result


def _is_empty(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def _infer_cell_type(v: Any) -> str:
    """单个值的粗类型推断。"""
    if _is_empty(v):
        return "empty"
    if isinstance(v, bool):
        return "bool"
    if isinstance(v, (int, float)):
        return "numeric"
    if isinstance(v, (datetime.datetime, datetime.date, datetime.time)):
        return "datetime"
    return "text"


def _infer_column_type(values: list) -> str:
    """根据列样本（前 N 行）推断候选类型。"""
    non_empty = [v for v in values if not _is_empty(v)]
    if not non_empty:
        return "empty"
    types = set(_infer_cell_type(v) for v in non_empty)
    if types <= {"numeric"}:
        return "numeric"
    if types <= {"datetime"}:
        return "datetime"
    if types <= {"bool"}:
        return "bool"
    if types <= {"numeric", "datetime"}:
        return "mixed"
    return "text"


def _cell_to_str(v: Any) -> str:
    """单元格值转展示字符串。"""
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float):
        # 避免浮点科学计数法；整数浮点去 .0
        if v == int(v):
            return str(int(v))
        return repr(v)
    if isinstance(v, bool):
        return str(v)
    return str(v).strip()


def _build_preview(rows: list, col_start: int, col_end: int,
                   merged_top_lefts: set) -> list[str]:
    """构建带行列号的 preview 文本列表：'行号| A:值 | B:值 | ...'。

    合并区域左上角单元格值前加 [合并] 标记，帮助大模型识别标题/表单区。
    """
    lines: list[str] = []
    for r_idx, row in enumerate(rows, start=1):
        cells: list[str] = []
        for c_idx in range(col_start, col_end + 1):
            v = row[c_idx - 1] if (c_idx - 1) < len(row) else None
            letter = _col_letter(c_idx)
            s = _cell_to_str(v)
            if (r_idx, c_idx) in merged_top_lefts:
                s = f"[合并]{s}"
            if len(s) > 80:
                s = s[:77] + "..."
            cells.append(f"{letter}:{s}")
        line = f"{r_idx}| " + " | ".join(cells)
        if len(line) > _MAX_ROW_CHARS:
            line = line[:_MAX_ROW_CHARS] + "...(行截断)"
        lines.append(line)
    return lines


def _build_column_hints(rows: list, col_start: int, col_end: int,
                        sample_size: int = 3) -> list[dict]:
    """每列的辅助摘要：非空率/候选类型/样本值。仅基于前 N 行。"""
    hints: list[dict] = []
    total = len(rows)
    for c_idx in range(col_start, col_end + 1):
        col_values = [row[c_idx - 1] if (c_idx - 1) < len(row) else None for row in rows]
        non_empty = [v for v in col_values if not _is_empty(v)]
        ratio = round(len(non_empty) / total, 2) if total else 0.0
        cand_type = _infer_column_type(col_values)
        samples = [_cell_to_str(v) for v in non_empty[:sample_size]]
        hints.append({
            "col": _col_letter(c_idx),
            "non_empty_ratio": ratio,
            "candidate_type": cand_type,
            "samples": samples,
        })
    return hints


def _suggest_data_region(rows: list, total_rows: int,
                        col_start: int, col_end: int) -> dict:
    """轻量建议：找首个"宽度稳定"的行作为 header_row 候选，data_end 用 total_rows。

    仅作参考，大模型可否决。判定：某行非空列>=3 且下一行也非空列>=3。
    data_end 默认用整表行数；若尾部有合计/制表人等行，大模型可将其调小。
    """
    n = len(rows)
    for i in range(n):
        row = rows[i]
        non_empty_cols = [
            c_idx for c_idx in range(col_start, col_end + 1)
            if (c_idx - 1) < len(row) and not _is_empty(row[c_idx - 1])
        ]
        if len(non_empty_cols) < 3:
            continue
        if i + 1 < n:
            next_row = rows[i + 1]
            next_non_empty = [
                c_idx for c_idx in range(col_start, col_end + 1)
                if (c_idx - 1) < len(next_row) and not _is_empty(next_row[c_idx - 1])
            ]
            if len(next_non_empty) >= 3:
                return {
                    "header_row": i + 1,  # 1-indexed
                    "data_start": i + 2,
                    "data_end": total_rows,  # 默认取整表行数，大模型可复核调整
                    "col_start": col_start,
                    "col_end": col_end,
                    "note": "仅参考（首个宽度稳定行）请大模型复核。data_end 默认用整表行数，"
                            "若预览中可见尾部合计/制表人等行请调小 data_end。",
                }
    # 未找到稳定行 → 仍返回含 data_end 的基础建议，供大模型自行判断
    return {
        "header_row": 1,
        "data_start": 1,
        "data_end": total_rows,
        "col_start": col_start,
        "col_end": col_end,
        "note": "未找到明显数据区，请大模型自行判断各边界（header_row/data_start/data_end）。",
    }


def _process_sheet(wb, sheet_name: str, preview_rows: int,
                   preview_cols: int | None, compact: bool,
                   merged_refs: list[str]) -> dict:
    """处理单个 sheet：读前 N 行，构建 preview + 摘要。"""
    ws = wb[sheet_name]
    actual_preview = min(preview_rows, 20) if compact else preview_rows

    # 流式读前 actual_preview 行（values_only 提高性能）
    rows: list = []
    for row in ws.iter_rows(min_row=1, max_row=actual_preview, values_only=True):
        rows.append(row)
        if len(rows) >= actual_preview:
            break

    # 维度（read_only 下 max_row/max_column 通常可用，可能不准，标注 estimated）
    try:
        total_rows = ws.max_row or len(rows)
    except Exception:
        total_rows = len(rows)
    try:
        total_cols = ws.max_column or (max((len(r) for r in rows), default=0))
    except Exception:
        total_cols = max((len(r) for r in rows), default=0)

    col_start = 1
    col_end = preview_cols if preview_cols else total_cols

    # 合并单元格（仅扫描窗内；从 zip 解析，read_only 模式 ws.merged_cells 为空）
    from openpyxl.utils import range_boundaries
    merged_top_lefts: set = set()
    merged_ranges: list[str] = []
    for ref in merged_refs:
        try:
            min_col, min_row, _max_col, _max_row = range_boundaries(ref)
            if min_row <= actual_preview:
                merged_ranges.append(ref)
                merged_top_lefts.add((min_row, min_col))
        except Exception:
            pass

    preview = _build_preview(rows, col_start, col_end, merged_top_lefts)
    column_hints = _build_column_hints(rows, col_start, col_end)
    suggested = _suggest_data_region(rows, total_rows, col_start, col_end)

    return {
        "name": sheet_name,
        "total_rows": total_rows,
        "total_cols": total_cols,
        "previewed_rows": len(rows),
        "preview": preview,
        "merged_ranges": merged_ranges,
        "column_hints": column_hints,
        "suggested_data_region": suggested,
    }


class HardExcelReadInput(BaseModel):
    file_path: str = Field(..., description="Excel 文件绝对路径")
    sheet: str | None = Field(
        None,
        description="指定单个 sheet 名；None=按 sheets 或全部 sheet 处理",
    )
    sheets: list[str] | None = Field(
        None,
        description="指定一批 sheet 名（分批用）；None 且 sheet 为空=全部 sheet",
    )
    preview_rows: int = Field(
        25,
        description="每个 sheet 读取的前 N 行（建议 25 或 50；所有 sheet 模式建议降到 15-20）",
    )
    preview_cols: int | None = Field(
        None,
        description="每行最多返回的列数；None=全部列",
    )
    all_sheets_compact: bool = Field(
        False,
        description="多 sheet 模式下是否用紧凑摘要(每 sheet 前 20 行)以控体积",
    )


class HardExcelReadTool(BaseTool):
    """读取 Excel 前 N 行 + 辅助摘要，供大模型解析复杂表结构。只读。"""

    args_schema: type[BaseModel] = HardExcelReadInput
    name: str = "hard_excel_read"
    description: str = (
        "复杂 Excel 表结构读取工具（hard_excel_read）。读取指定 sheet（或所有 sheet）的前 N 行原始数据"
        "+ 辅助摘要（维度/合并区/列类型与样本/数据区范围建议），返回给大模型解析复杂表结构。\n"
        "适用场景：Excel 表结构复杂（含标题区/表单说明区/行列式数据区/尾部区），需要先理解结构再入库。\n"
        "本工具不内部调大模型，只返回前 N 行原始数据与摘要；由大模型据此产出 table_description 和 data_region，供 excel2sql 使用。\n"
        "绝不返回数据区全量行，只返回前 N 行，控制上下文体积。\n"
        "参数说明：\n"
        "- file_path: Excel 文件绝对路径\n"
        "- sheet: 指定单个 sheet 名；None=按 sheets 或全部 sheet\n"
        "- sheets: 指定一批 sheet 名（分批）；None 且 sheet 为空=全部 sheet\n"
        "- preview_rows: 每 sheet 读取前 N 行（建议 25 或 50；所有 sheet 模式建议 15-20）\n"
        "- preview_cols: 每行最多列数；None=全部\n"
        "- all_sheets_compact: 多 sheet 时用紧凑摘要(每 sheet 前 20 行)控体积\n"
        "权限：只读，direct 直接执行。\n"
        "调用示例：\n"
        '- 单 sheet 深看: {"file_path": "C:\\\\data\\\\sales.xlsx", "sheet": "2023年销售明细", "preview_rows": 25}\n'
        '- 所有 sheet 摸底: {"file_path": "C:\\\\data\\\\sales.xlsx", "preview_rows": 20, "all_sheets_compact": true}\n'
    )

    def _run(
        self,
        file_path: str,
        sheet: str | None = None,
        sheets: list[str] | None = None,
        preview_rows: int = 25,
        preview_cols: int | None = None,
        all_sheets_compact: bool = False,
        **kwargs: Any,
    ) -> str:
        try:
            if not os.path.isfile(file_path):
                return json.dumps({"error": f"文件不存在: {file_path}"}, ensure_ascii=False)
            if preview_rows < 1:
                preview_rows = 25

            # 按文件格式选择引擎
            if _file_is_xls(file_path):
                return self._run_xls(file_path, sheet, sheets, preview_rows, preview_cols, all_sheets_compact)

            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            try:
                available = list(wb.sheetnames)

                # 确定要处理的 sheet 列表
                if sheet:
                    target_sheets = [sheet]
                elif sheets:
                    target_sheets = list(sheets)
                else:
                    target_sheets = available

                # 校验 sheet 存在
                avail_set = set(available)
                missing = [s for s in target_sheets if s not in avail_set]
                if missing:
                    return json.dumps(
                        {"error": f"sheet 不存在: {missing}", "available_sheets": available},
                        ensure_ascii=False,
                    )

                compact = all_sheets_compact and len(target_sheets) > 1

                # 合并单元格从 zip 解析（read_only 模式 ws.merged_cells 为空）
                merged_map = _read_all_merged_cells(file_path)

                result_sheets: list[dict] = []
                truncated = False
                for sn in target_sheets:
                    sheet_data = _process_sheet(wb, sn, preview_rows, preview_cols, compact, merged_map.get(sn, []))
                    result_sheets.append(sheet_data)
                    # 体积保护：超限则停止并提示分批
                    probe = json.dumps(result_sheets, ensure_ascii=False, default=str)
                    if len(probe) > _MAX_RETURN_CHARS:
                        truncated = True
                        result_sheets.pop()  # 回退最后一个超额 sheet
                        break

                mode = "single_sheet" if len(target_sheets) == 1 else "multi_sheet"
                result: dict = {
                    "workbook": file_path,
                    "mode": mode,
                    "total_sheets_in_workbook": len(available),
                    "processed_sheets": len(result_sheets),
                    "sheets": result_sheets,
                    "note": (
                        "请基于每 sheet 的 preview 原始数据与辅助摘要，判断："
                        "(1) table_description：数据区上方的标题/表单/说明等复杂信息，整合成一段文本；"
                        "(2) data_region：真正的行列式数据区范围 "
                        "{header_row, data_start, data_end, col_start, col_end}（1-indexed）。"
                        "结果供 excel2sql 工具使用。"
                    ),
                }
                if truncated:
                    processed_names = [s["name"] for s in result_sheets]
                    remaining = [s for s in target_sheets if s not in processed_names]
                    result["truncated"] = True
                    result["truncated_hint"] = (
                        f"返回体积超限已截断。已处理: {processed_names}；"
                        f"请用 sheets 参数分批处理剩余: {remaining}"
                    )
                return json.dumps(result, ensure_ascii=False, default=str)
            finally:
                wb.close()
        except Exception as e:
            return json.dumps({"error": f"hard_excel_read 失败: {str(e)}"}, ensure_ascii=False)

    def _process_sheet_xls(self, book, sheet_name: str, preview_rows: int,
                           preview_cols: int | None, compact: bool) -> dict:
        """xlrd 版：处理单个 sheet。"""
        sheet = book.sheet_by_name(sheet_name)
        actual_preview = min(preview_rows, 20) if compact else preview_rows
        actual_preview = min(actual_preview, sheet.nrows)

        total_rows = sheet.nrows
        total_cols = sheet.ncols
        col_end = min(preview_cols or total_cols, total_cols)

        # 读前 N 行
        rows: list = []
        for r in range(actual_preview):
            row_vals = [_xlrd_cell_to_value(sheet, r, c) if c < total_cols else None
                        for c in range(col_end)]
            rows.append(row_vals)

        merged_ranges, merged_top_lefts = _get_xls_merged(sheet, actual_preview)
        preview = _build_preview(rows, 1, col_end, merged_top_lefts)
        column_hints = _build_column_hints(rows, 1, col_end)
        suggested = _suggest_data_region(rows, total_rows, 1, col_end)

        return {
            "name": sheet_name,
            "total_rows": total_rows,
            "total_cols": total_cols,
            "previewed_rows": len(rows),
            "preview": preview,
            "merged_ranges": merged_ranges,
            "column_hints": column_hints,
            "suggested_data_region": suggested,
        }

    def _run_xls(
        self,
        file_path: str,
        sheet: str | None,
        sheets: list[str] | None,
        preview_rows: int,
        preview_cols: int | None,
        all_sheets_compact: bool,
    ) -> str:
        """xlrd 引擎：处理 .xls 文件。"""
        book, _ = _load_workbook_xls(file_path)
        available = book.sheet_names()

        if sheet:
            target_sheets = [sheet]
        elif sheets:
            target_sheets = list(sheets)
        else:
            target_sheets = available

        avail_set = set(available)
        missing = [s for s in target_sheets if s not in avail_set]
        if missing:
            return json.dumps(
                {"error": f"sheet 不存在: {missing}", "available_sheets": available},
                ensure_ascii=False,
            )

        compact = all_sheets_compact and len(target_sheets) > 1

        result_sheets: list[dict] = []
        truncated = False
        for sn in target_sheets:
            sheet_data = self._process_sheet_xls(book, sn, preview_rows, preview_cols, compact)
            result_sheets.append(sheet_data)
            probe = json.dumps(result_sheets, ensure_ascii=False, default=str)
            if len(probe) > _MAX_RETURN_CHARS:
                truncated = True
                result_sheets.pop()
                break

        mode = "single_sheet" if len(target_sheets) == 1 else "multi_sheet"
        result: dict = {
            "workbook": file_path,
            "mode": mode,
            "engine": "xlrd",
            "total_sheets_in_workbook": len(available),
            "processed_sheets": len(result_sheets),
            "sheets": result_sheets,
            "note": (
                "请基于每 sheet 的 preview 原始数据与辅助摘要，判断："
                "(1) table_description：数据区上方的标题/表单/说明等复杂信息，整合成一段文本；"
                "(2) data_region：真正的行列式数据区范围 "
                "{header_row, data_start, data_end, col_start, col_end}（1-indexed）。"
                "结果供 excel2sql 工具使用。"
            ),
        }
        if truncated:
            processed_names = [s["name"] for s in result_sheets]
            remaining = [s for s in target_sheets if s not in processed_names]
            result["truncated"] = True
            result["truncated_hint"] = (
                f"返回体积超限已截断。已处理: {processed_names}；"
                f"请用 sheets 参数分批处理剩余: {remaining}"
            )
        return json.dumps(result, ensure_ascii=False, default=str)

    async def _arun(self, **kwargs: Any) -> str:
        import asyncio
        return await asyncio.to_thread(self._run, **kwargs)
