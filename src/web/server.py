# server.py
"""
为 ``web`` 目录下的静态前端提供 API：会话列表、多模态发送（与 ``agent_runtime`` 一致）。

从仓库 ``Agent`` 目录、且 ``PYTHONPATH`` 包含 ``src`` 时启动::

    uvicorn agent.web.server:app --reload --host 127.0.0.1 --port 8765
"""
from __future__ import annotations

import asyncio
import base64
import json
import mimetypes
import os
import queue as _queue_module
import shutil
import sys
import tempfile
import uuid
from contextlib import asynccontextmanager
from pathlib import Path
from typing import Any

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
_SRC = Path(__file__).resolve().parent.parent.parent
if str(_SRC) not in sys.path:
    sys.path.insert(0, str(_SRC))

from agent.core import turn_runner  # noqa: E402
from web import ui_session as ui  # noqa: E402
from agent.utils.image_utils import IMAGE_FILE_EXTENSIONS  # noqa: E402
from web.ui_session import _AUDIO_EXTENSIONS  # noqa: E402
from agent.core.config_manager import (  # noqa: E402
    load_agent_configs, save_agent_configs,
    load_tool_configs, save_tool_configs,
    load_env_config, save_env_config,
    load_models, save_models,
    get_registered_tool_names,
    load_gui_config, save_gui_config,
    _get_live_monitors,
    AGENT_CONFIG_PATH, TOOL_CONFIG_PATH, ENV_CONFIG_PATH,
    GUI_CONFIG_PATH,
)

_WEB_DIR = Path(__file__).resolve().parent
_IMAGE_DIR = _WEB_DIR / "image"
_CONFIG_DIR = _WEB_DIR.parent / "agent" / "config"  # src/agent/config/

_AUDIO_EXT = frozenset(_AUDIO_EXTENSIONS)

# ---------- 工作空间配置持久化 ----------

_WS_CONFIG_PATH = _CONFIG_DIR / "workspace_config.json"


def _read_ws_config() -> dict:
    """读取工作空间配置文件。"""
    if _WS_CONFIG_PATH.is_file():
        try:
            with open(_WS_CONFIG_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            pass
    return {"current": "", "list": [], "access_mode": "restricted"}


def _write_ws_config(data: dict) -> bool:
    """写入工作空间配置文件。"""
    try:
        with open(_WS_CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except OSError:
        return False


# 启动时加载已保存的工作空间配置
def _init_workspace():
    """从 workspace_config.json 恢复上次选择的工作空间。"""
    cfg = _read_ws_config()
    current = cfg.get("current", "")
    ws_list = cfg.get("list", [])
    changed = False

    # 清理列表中已不存在的文件夹
    valid = [p for p in ws_list if os.path.isdir(p)]
    if len(valid) != len(ws_list):
        cfg["list"] = valid
        changed = True

    # 上次选择的工作空间已不存在时，回退到默认工作空间
    if current and not os.path.isdir(current):
        current = ""
        cfg["current"] = ""
        changed = True

    if changed:
        _write_ws_config(cfg)
    if current:
        from agent.memory.system_prompt import set_current_workspace
        set_current_workspace(current)
        # 同步 env_utils.WORKSPACE_DIR
        try:
            import agent.utils.env_utils as env_utils
            env_utils.WORKSPACE_DIR = current
        except Exception:
            pass

_init_workspace()


def _clean_tts_text(text: str) -> str:
    """清理 TTS 输入文本：去掉表情/特殊符号，按句号换行帮助分句。"""
    import re
    # 保留：中文、英文、数字、常用标点
    allowed = re.compile(
        r'[\u4e00-\u9fff'           # 中文
        r'a-zA-Z0-9'                # 英文数字
        r'，。！？；：、""''（）《》'  # 中文标点
        r'.,!?;:()\[\]{}<>/\-\u2014\u2015\u2026\u00b7'  # 英文标点+破折号省略号
        r'\s'                       # 空白
        r']'
    )
    cleaned = ''.join(ch if allowed.match(ch) else ' ' for ch in text)
    # 合并多余空白
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    # 按句号/感叹号/问号换行
    cleaned = re.sub(r'([。！？!?])', r'\1\n', cleaned)
    # 去掉空行
    lines = [l.strip() for l in cleaned.split('\n') if l.strip()]
    return '\n'.join(lines)


def _ensure_session_in_list(session_id: str, ids: list[str]) -> list[str]:
    """确保 session_id 在会话 ID 列表中（置顶）。"""
    if session_id not in ids:
        ids = [session_id, *ids]
    return ids


def _data_url_for_file(fp: str) -> str | None:
    """将本地文件读入为 data URL，供前端直接展示（不单独提供文件下载路由）。

    输入:
        fp: 本地文件路径。

    输出:
        data URL 字符串，失败返回 None。

    系统定位:
        _serialize_content 将附件转为内联 data URL 以嵌入 API 响应。

    可扩展性:
        可增加缓存、大小限制、权限检查。
    """
    if not fp or not os.path.isfile(fp):
        return None
    mime, _ = mimetypes.guess_type(fp)
    if not mime:
        ext = os.path.splitext(fp)[1].lower()
        if ext in IMAGE_FILE_EXTENSIONS:
            mime = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
                    ".gif": "image/gif", ".webp": "image/webp", ".bmp": "image/bmp",
                    ".ico": "image/x-icon",
                    ".svg": "image/svg+xml"}.get(ext, "image/jpeg")
        elif ext in _AUDIO_EXT:
            mime = {
                ".wav": "audio/wav",
                ".mp3": "audio/mpeg",
                ".m4a": "audio/mp4",
                ".flac": "audio/flac",
                ".ogg": "audio/ogg",
            }.get(ext, "application/octet-stream")
        else:
            mime = "application/octet-stream"
    try:
        with open(fp, "rb") as f:
            raw = f.read()
        b64 = base64.standard_b64encode(raw).decode("ascii")
    except OSError:
        return None
    return f"data:{mime};base64,{b64}"


def _serialize_content(content: Any, *, meta: dict[str, Any] | None = None, role: str | None = None) -> Any:
    """将一条消息中的 content 转为 JSON 友好结构；文件以内联 data URL 输出。

    输入:
        content: 原始 content（str、list 或 dict）。
        meta: 消息的 meta 信息（如音频时长、自动播放等）。
        role: 消息角色（user/assistant），用于控制自动播放行为。

    输出:
        序列化后的 content（str 或 list of dict）。

    系统定位:
        _serialize_messages 中逐条转换，供前端直接渲染。

    可扩展性:
        可支持更多文件类型（如视频）的 data URL 输出。
    """
    meta = meta or {}
    if isinstance(content, str):
        return content
    if not isinstance(content, list):
        return str(content)
    parts: list[dict[str, Any]] = []
    for piece in content:
        if isinstance(piece, str) and piece.strip():
            parts.append({"type": "text", "text": piece.strip()})
            continue
        if not isinstance(piece, dict):
            continue
        if piece.get("type") == "text":
            parts.append({"type": "text", "text": piece.get("text") or ""})
            continue
        fp = piece.get("path") or (piece.get("file") or {}).get("path")
        display_name = piece.get("name")
        is_typed_audio = piece.get("type") == "audio"
        if piece.get("type") == "image_url":
            fp = fp or (piece.get("image_url") or {}).get("url")
        if is_typed_audio:
            fp = fp or piece.get("audio_path")
        if not fp or not os.path.isfile(fp):
            continue
        data_url = _data_url_for_file(fp)
        if not data_url:
            continue
        ext = os.path.splitext(fp)[1].lower()
        if ext in IMAGE_FILE_EXTENSIONS:
            parts.append({"type": "image", "url": data_url, "path": fp, "name": display_name or os.path.basename(fp)})
        elif is_typed_audio and ext in _AUDIO_EXT:
            # 仅助手 TTS 语音回复（piece.type == "audio"）保留 audio part 与 duration；
            # 用户上传的音频文件（无 type 或 type != "audio"）按普通文件展示。
            audio_meta = meta.get("audio") if isinstance(meta.get("audio"), dict) else {}
            parts.append({
                "type": "audio",
                "url": data_url,
                "path": fp,
                "name": display_name or os.path.basename(fp),
                "autoplay": bool(meta.get("autoplay") or audio_meta.get("autoplay")) if role == "assistant" else False,
                "duration_seconds": audio_meta.get("duration_seconds"),
            })
        else:
            parts.append({"type": "file", "url": data_url, "path": fp, "name": display_name or os.path.basename(fp)})
    return parts


def _serialize_messages(hist: list[dict[str, Any]] | None, session_id: str = "") -> list[dict[str, Any]]:
    """将 chat_history 转为 API 响应格式，附件内联为 data URL。

    输入:
        hist: 聊天历史（dict 列表）。
        session_id: 会话 ID，用于附加工具调用历史。

    输出:
        序列化后的消息列表，每条消息包含 role、content、meta。

    系统定位:
        所有返回给前端的 API 响应中，messages 字段均经过此函数转换。

    可扩展性:
        可支持更细粒度的权限控制或裁剪。
    """
    hist = hist or []
    # 找到最后一条 assistant 消息的索引，用于控制自动播放
    last_assistant_idx = -1
    for i in range(len(hist) - 1, -1, -1):
        if hist[i].get("role") == "assistant":
            last_assistant_idx = i
            break

    out = []
    for idx, m in enumerate(hist):
        role = m.get("role")
        content = m.get("content")
        meta = dict(m.get("meta") or {})
        meta.setdefault("output_type", "text" if role != "assistant" else meta.get("output_type") or "text")

        serialized_content = _serialize_content(content, meta=meta, role=role)

        out.append({"role": role, "content": serialized_content, "meta": meta})

    # 附加工具调用历史到消息（供前端持久化展示）
    if session_id and out:
        try:
            from agent.history.tool_call_recorder import (
                get_tool_calls_for_turn, get_tool_calls_for_latest_turn,
                get_sub_agent_traces_for_turn, get_turn_meta,
            )
        except Exception:
            return out

        # 按 turn_id 为每条 assistant 消息挂载对应轮次的工具调用与轮次级元数据
        has_turn_ids = False
        for m in out:
            if m.get("role") != "assistant":
                continue
            turn_id = (m.get("meta") or {}).get("turn_id", "")
            if not turn_id:
                continue
            has_turn_ids = True
            m_meta = m.setdefault("meta", {})
            try:
                tool_calls = get_tool_calls_for_turn(session_id, turn_id)
            except Exception:
                tool_calls = []
            if tool_calls:
                _inject_sub_tool_calls(tool_calls, session_id, turn_id, get_sub_agent_traces_for_turn)
                m_meta["tool_calls"] = tool_calls
            # 无条件挂载 turn_meta（started_at/ended_at/duration），供前端显示回复开始时间
            try:
                turn_meta = get_turn_meta(session_id, turn_id)
            except Exception:
                turn_meta = None
            if turn_meta:
                m_meta["turn_meta"] = turn_meta

        # 兼容旧消息（无 turn_id）：将最新轮工具调用附加到最后一条 assistant
        if not has_turn_ids:
            try:
                tool_calls = get_tool_calls_for_latest_turn(session_id)
            except Exception:
                tool_calls = []
            if tool_calls:
                _inject_sub_tool_calls(tool_calls, session_id, "", get_sub_agent_traces_for_turn)
                from agent.history.tool_call_recorder import get_latest_turn
                latest_rec = get_latest_turn(session_id)
                turn_meta = {}
                if latest_rec:
                    turn_meta = {
                        k: latest_rec.get(k)
                        for k in ("started_at", "ended_at", "duration")
                        if latest_rec.get(k) is not None
                    }
                for m in reversed(out):
                    if m.get("role") == "assistant":
                        m_meta = m.setdefault("meta", {})
                        m_meta["tool_calls"] = tool_calls
                        if turn_meta:
                            m_meta["turn_meta"] = turn_meta
                        break

    return out


def _inject_sub_tool_calls(
    tool_calls: list[dict], session_id: str, turn_id: str,
    lookup_fn,
) -> None:
    """为 tool_calls 中的 call_subagent 条目注入嵌套的子 Agent 轨迹。

    Args:
        tool_calls: 工具调用列表（原地修改）。
        session_id: 会话 ID。
        turn_id: 轮次 ID。
        lookup_fn: (session_id, turn_id, agent_name, tool_call_id) -> list[dict] 的查询函数。
    """
    for tc in tool_calls:
        if tc.get("name") != "call_subagent" and tc.get("name") != "agent_call":
            continue
        # 已经携带子轨迹则跳过（已从持久化 JSON 中加载）
        if tc.get("sub_tool_calls"):
            continue
        args = tc.get("args", {})
        if isinstance(args, str):
            try:
                import json
                args = json.loads(args)
            except (json.JSONDecodeError, TypeError):
                args = {}
        sub_name = args.get("agent_name", "")
        if sub_name:
            # 优先按条目携带的 tool_call_id 精确匹配（并行/同名子 Agent 各自独立），
            # lookup_fn 内部对空 tool_call_id 回退到 agent_name 级（旧数据兼容）。
            sub_tc = lookup_fn(session_id, turn_id, sub_name, tc.get("id", ""))
            if sub_tc:
                tc["sub_tool_calls"] = sub_tc


def _bootstrap() -> tuple[str, list[str], list[dict[str, Any]]]:
    """初始化默认会话：取最近会话或新建时间戳 ID 及其消息列表。

    输入: 无。

    输出:
        (session_id, sessions列表, messages列表)

    系统定位:
        API 启动时/删除会话后回退到默认会话。

    可扩展性:
        可增加用户偏好默认会话逻辑。
    """
    ids = ui.list_session_ids_ordered()
    if ids:
        sid = ids[0]
        hist = ui.load_ui_messages_for_session(sid)
    else:
        sid = ui.new_timestamp_session_id()
        hist = []
        ids = [sid]
    return sid, ids, hist


@asynccontextmanager
async def lifespan(app):
    """应用生命周期：预热重依赖 + 启动定时任务调度器，退出时关闭。

    预热目的:
        agent.tools（含 chromadb 等重依赖）与 Agent 运行时原先在首次
        工具调用/首次打开设置页时惰性导入，首次请求要现场付数秒导入成本。
        这里在启动阶段统一预热（成本被 Electron 健康检查吸收），
        使设置页与首次聊天即时响应。
    """
    try:
        # 预热：构建全部 Agent 运行时（顺带导入 agent.tools/chromadb 等重依赖）
        from agent.agents import agent_runtime  # noqa: F401
        # 预热工具参数缓存（实例化全部工具类读取 args_schema）
        from agent.core.config_manager import _get_tool_parameters
        _get_tool_parameters()
    except Exception as e:
        print(f"[lifespan] 预热 Agent 运行时失败: {e}")
    try:
        from agent.cron.scheduler import get_scheduler
        get_scheduler().start()
    except Exception as e:
        print(f"[lifespan] 启动 CronScheduler 失败: {e}")
    yield
    try:
        from agent.cron.scheduler import get_scheduler
        get_scheduler().shutdown()
    except Exception:
        pass


app = FastAPI(title="minor web chat", version="1.0.0", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 禁用缓存 — Electron 端每次加载都是从 FastAPI 取最新文件
@app.middleware("http")
async def no_cache_middleware(request, call_next):
    response = await call_next(request)
    if request.url.path.startswith("/api/"):
        return response  # API 路由不干预
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


@app.get("/api/bootstrap")
def api_bootstrap() -> dict[str, Any]:
    """获取初始会话信息（默认会话 ID、会话列表、消息历史、当前 token 用量）。

    输入: 无。

    输出:
        {"sessionId": str, "sessions": list[str], "messages": list[dict], "tokens": int}

    系统定位:
        前端页面加载时调用的第一个 API，用于初始化界面。

    可扩展性:
        可返回用户配置、主题等元信息。
    """
    sid, ids, hist = _bootstrap()
    from agent.history.tool_call_recorder import get_session_tokens
    return {
        "sessionId": sid,
        "sessions": ids,
        "messages": _serialize_messages(hist, sid),
        "tokens": get_session_tokens(sid),
    }


@app.post("/api/sessions/new")
def api_new_session() -> dict[str, Any]:
    """创建新会话（基于时间戳），返回会话信息。

    输入: 无。

    输出:
        {"sessionId": str, "sessions": list[str], "messages": []}

    系统定位:
        前端点击“新建会话”时调用。
    """
    sid = ui.new_timestamp_session_id()
    ids = ui.list_session_ids_ordered()
    ids = [sid] + [x for x in ids if x != sid]
    return {"sessionId": sid, "sessions": ids, "messages": [], "tokens": 0}


def _branch_copy_attachment(abs_src: str, dst_dir: str, new_sid: str) -> str:
    """将附件复制到新会话目录，返回目标绝对路径（必要时避免重名）。

    输入:
        abs_src: 源附件绝对路径。
        dst_dir: 新会话目录。
        new_sid: 新会话 ID（用于冲突重命名）。

    输出:
        目标绝对路径。
    """
    base = os.path.basename(abs_src)
    dest = os.path.join(dst_dir, base)
    if os.path.exists(dest) and os.path.normpath(dest) != os.path.normpath(abs_src):
        stem, ext = os.path.splitext(base)
        i = 2
        while os.path.exists(dest):
            dest = os.path.join(dst_dir, f"{stem}_{new_sid}_{i}{ext}")
            i += 1
    if os.path.normpath(dest) != os.path.normpath(abs_src):
        shutil.copy2(abs_src, dest)
    return dest


def _branch_rewrite_attachments(msgs: list[dict], dst_dir: str, new_sid: str) -> None:
    """将分支消息中的附件引用复制到新会话目录并重写 relpath，使新会话自包含。

    输入:
        msgs: 分支消息列表（user 条目 relpath / assistant 条目 audio_relpath）。
        dst_dir: 新会话目录。
        new_sid: 新会话 ID（用于冲突重命名）。

    系统定位:
        附件二进制文件始终落盘（两个后端一致），此处仅处理消息内容中的
        附件引用重写，与存储后端无关。
    """
    from agent.utils.agent_utils import SESSIONS_ROOT

    for msg in msgs:
        role = msg.get("role")
        if role == "user":
            content = (msg.get("user") or {}).get("content") or []
            for piece in content:
                if not isinstance(piece, dict):
                    continue
                rel = piece.get("relpath")
                if not rel:
                    continue
                abs_src = os.path.normpath(os.path.join(SESSIONS_ROOT, rel))
                if not os.path.isfile(abs_src):
                    continue
                dest = _branch_copy_attachment(abs_src, dst_dir, new_sid)
                new_rel = os.path.relpath(dest, SESSIONS_ROOT).replace("\\", "/")
                if new_rel != rel:
                    piece["relpath"] = new_rel
        elif role == "assistant":
            asst = msg.get("assistant") or {}
            audio_rel = asst.get("audio_relpath")
            if audio_rel:
                abs_src = os.path.normpath(os.path.join(SESSIONS_ROOT, audio_rel.replace("/", os.sep)))
                if os.path.isfile(abs_src):
                    dest = _branch_copy_attachment(abs_src, dst_dir, new_sid)
                    new_rel = os.path.relpath(dest, SESSIONS_ROOT).replace("\\", "/")
                    if new_rel != audio_rel:
                        asst["audio_relpath"] = new_rel


@app.post("/api/sessions/branch")
def api_session_branch(
    session_id: str = Form(...),
    branch_turn_id: str = Form(""),
) -> dict[str, Any]:
    """将源会话中 branch_turn_id 及之前的所有历史复制到新会话（分支）。

    输入:
        session_id: 源会话 ID。
        branch_turn_id: 分支点回合 ID（该回合及其之前的历史会被复制到新会话）。

    输出:
        {"sessionId": str, "sessions": list[str], "messages": list[dict], "tokens": int}

    系统定位:
        AI 回复的「分支」按钮：复制一份历史到新会话，用户可在此基础上继续对话。
        基于统一存储 API 实现（消息/工具记录读写均后端无关，mysql 后端同样支持）。
    """
    if not session_id:
        raise HTTPException(status_code=400, detail="缺少 session_id")

    from agent.utils.agent_utils import SESSIONS_ROOT
    from agent.history import session_storage

    # 源会话 turn 列表（后端无关）
    turn_ids = session_storage.list_turn_ids(session_id)
    if not turn_ids:
        raise HTTPException(status_code=400, detail="源会话不存在或为空")
    branch_ids = [t for t in turn_ids if (not branch_turn_id) or t <= branch_turn_id]
    if not branch_ids:
        raise HTTPException(status_code=400, detail="无效的分支回合")

    # 创建新会话并复制 branch_turn_id 及之前的消息与工具调用记录
    new_sid = ui.new_timestamp_session_id()
    dst_dir = os.path.join(SESSIONS_ROOT, new_sid)
    os.makedirs(dst_dir, exist_ok=True)
    try:
        for tid in branch_ids:
            msgs = session_storage.load_turn_messages(session_id, tid) or []
            # 复制附件引用并重写 relpath，使新会话自包含
            _branch_rewrite_attachments(msgs, dst_dir, new_sid)
            session_storage.save_turn(new_sid, tid, msgs)
            rec = session_storage.get_turn_record(session_id, tid)
            if rec:
                _meta = {k: rec[k] for k in ("started_at", "ended_at", "duration", "aborted") if k in rec}
                session_storage.save_tool_calls(
                    new_sid, tid, rec.get("tool_calls") or [], meta=_meta or None
                )
    except Exception:
        shutil.rmtree(dst_dir, ignore_errors=True)
        raise HTTPException(status_code=400, detail="无效的分支回合")

    ids = _ensure_session_in_list(new_sid, ui.list_session_ids_ordered())
    from agent.history.tool_call_recorder import get_session_tokens
    return {
        "sessionId": new_sid,
        "sessions": ids,
        "messages": _serialize_messages(ui.load_ui_messages_for_session(new_sid), new_sid),
        "tokens": get_session_tokens(new_sid),
    }


@app.get("/api/sessions/{session_id}/messages")
def api_session_messages(session_id: str) -> dict[str, Any]:
    """切换会话：加载指定会话的消息列表与全局会话列表。

    输入:
        session_id: 路径参数，目标会话 ID。

    输出:
        {"sessionId": str, "sessions": list[str], "messages": list[dict]}

    系统定位:
        前端切换会话时调用。
    """
    if not session_id:
        raise HTTPException(status_code=400, detail="缺少 session_id")
    hist = ui.load_ui_messages_for_session(session_id)
    ids = _ensure_session_in_list(session_id, ui.list_session_ids_ordered())
    from agent.history.tool_call_recorder import get_session_tokens
    return {
        "sessionId": session_id,
        "sessions": ids,
        "messages": _serialize_messages(hist, session_id),
        "tokens": get_session_tokens(session_id),
    }


@app.post("/api/sessions/clear")
def api_clear_session(session_id: str = Form(...)) -> dict[str, Any]:
    """清空当前会话的所有消息（重置会话）。

    输入:
        session_id: 表单字段，会话 ID。

    输出:
        {"sessionId": str, "sessions": list[str], "messages": []}

    系统定位:
        前端"清空对话"按钮调用。
    """
    if not session_id:
        raise HTTPException(status_code=400, detail="缺少 session_id")
    messages, _ = ui.clear_ui_and_turns(session_id)
    # 清理 tool_calling 目录
    _clear_tool_calling(session_id)
    # 清理 token 记录
    from agent.history.tool_call_recorder import clear_session_tokens
    clear_session_tokens(session_id)
    ids = _ensure_session_in_list(session_id, ui.list_session_ids_ordered())
    return {"sessionId": session_id, "sessions": ids, "messages": _serialize_messages(messages, session_id), "tokens": 0}


@app.post("/api/sessions/delete")
def api_delete_session(session_id: str = Form(...)) -> dict[str, Any]:
    """删除整个会话（含文件与历史），并回退到默认会话。

    输入:
        session_id: 表单字段，会话 ID。

    输出:
        {"sessionId": str, "sessions": list[str], "messages": list[dict]}

    系统定位:
        前端“删除会话”按钮调用。
    """
    if not session_id:
        raise HTTPException(status_code=400, detail="缺少 session_id")
    # 先设置 abort 标志并应答等待中的人工请求，通知正在执行的 Agent 停止
    try:
        from agent.core.runtime import set_abort_flag
        set_abort_flag(session_id)
    except Exception:
        pass
    try:
        from agent.core.human_request import resolve_all_for_session
        resolve_all_for_session(session_id, decision="reject", instruction="会话已删除")
    except Exception:
        pass
    ui.delete_session(session_id)
    turn_runner.drop_turn(session_id)
    # delete_session 已清理 tool_calling 目录，此处仅需清理 token
    try:
        from agent.history.tool_call_recorder import clear_session_tokens
        clear_session_tokens(session_id)
    except Exception:
        pass
    sid, ids, hist = _bootstrap()
    from agent.history.tool_call_recorder import get_session_tokens
    return {
        "sessionId": sid,
        "sessions": ids,
        "messages": _serialize_messages(hist, sid),
        "tokens": get_session_tokens(sid),
    }


def _materialize_uploads_to_temp(
    images: list[UploadFile] | None,
    files: list[UploadFile] | None,
    text: str,
) -> tuple[dict[str, Any], str]:
    """将上传文件落盘到临时目录，返回 ``(query_dict, tmp_root)``；调用方须在 finally 中删除 ``tmp_root``。

    输入:
        images: 前端上传的图片文件列表。
        files: 前端上传的普通文件列表（含音频文件，作为附件展示，不做 ASR）。
        text: 用户输入的文本。

    输出:
        (query_dict, tmp_root)
        query_dict 包含 "text", "files"(路径列表), "file_names"(原始文件名列表)。

    系统定位:
        API 层处理用户多模态输入的统一临时落盘逻辑。
        注意：语音输入（麦克风录音）由 ``voice_input`` 字段单独处理并执行 ASR 转
        文本，不再作为音频附件保存；上传的音频文件仍走 ``files`` 当普通文件展示。
    """
    tmp_root = tempfile.mkdtemp(prefix="web_mm_")
    file_paths: list[str] = []
    file_names: list[str] = []
    for uf in images or []:
        if uf is None or not uf.filename:
            continue
        ext = Path(uf.filename).suffix or ""
        dest = os.path.join(tmp_root, f"{uuid.uuid4().hex}{ext}")
        with open(dest, "wb") as f:
            shutil.copyfileobj(uf.file, f)
        file_paths.append(dest)
        file_names.append(uf.filename)

    for uf in files or []:
        if uf is None or not uf.filename:
            continue
        ext = Path(uf.filename).suffix or ""
        dest = os.path.join(tmp_root, f"{uuid.uuid4().hex}{ext}")
        with open(dest, "wb") as f:
            shutil.copyfileobj(uf.file, f)
        file_paths.append(dest)
        file_names.append(uf.filename)

    query = {"text": (text or "").strip(), "files": file_paths, "file_names": file_names}
    return query, tmp_root


def _save_voice_input_to_temp(voice_input: UploadFile | None) -> tuple[str | None, str]:
    """将语音输入文件落盘到临时目录，返回 ``(wav_path, tmp_root)``。

    输入:
        voice_input: 前端麦克风录音上传的音频文件。

    输出:
        (wav_path, tmp_root)
        wav_path 为落盘后的音频文件路径（供 ASR 读取）；无文件时为 (None, tmp_root)。
        调用方须在 finally 中删除 ``tmp_root``。

    系统定位:
        /api/chat/start 中处理语音输入的临时落盘逻辑，ASR 完成后即丢弃，不持久化。
    """
    tmp_root = tempfile.mkdtemp(prefix="web_voice_")
    if voice_input is None or not voice_input.filename:
        return None, tmp_root
    ext = Path(voice_input.filename).suffix or ".bin"
    dest = os.path.join(tmp_root, f"{uuid.uuid4().hex}{ext}")
    with open(dest, "wb") as f:
        shutil.copyfileobj(voice_input.file, f)
    return dest, tmp_root


def _transcribe_voice_input(voice_path: str) -> str:
    """调用 ASR 服务将语音输入转为文本。

    输入:
        voice_path: 临时音频文件路径。

    输出:
        识别出的文本字符串；ASR 失败时抛出 HTTPException。

    系统定位:
        /api/chat/start 阻塞调用，将用户语音输入转为文字后并入 user 消息 text。
    """
    from agent.utils.agent_utils import audio_file_to_text
    try:
        return audio_file_to_text(voice_path).strip()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"语音转文字失败: {e}")


def _chat_start_core(session_id: str, query: dict[str, Any]) -> dict[str, Any]:
    """追加 user、生成 turn_id、复制附件到 turn 目录（不调用模型）。

    输入:
        session_id: 会话 ID。
        query: {"text": str, "files": list[str], "file_names": list[str]}

    输出:
        {"sessionId": str, "sessions": list[str], "messages": list[dict]}
    """
    from datetime import datetime

    hist = ui.load_ui_messages_for_session(session_id)
    before = len(hist or [])
    new_hist = ui.add_message(hist or [], query)
    if new_hist is None:
        raise HTTPException(status_code=400, detail="请输入文字或上传图片/音频")
    if len(new_hist) == before:
        raise HTTPException(status_code=400, detail="请输入文字或上传图片/音频")

    # 生成 turn_id 并创建 turn 目录
    turn_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    ui.set_pending_turn_id(session_id, turn_id)

    # 先将用户附件复制到 turn 目录，更新 new_hist 中的路径
    ui.rewrite_user_files_for_last_turn(new_hist, session_id)

    # 再保存 turn 消息（此时路径已指向会话目录，而非临时目录）
    from web import ui_session as _ui
    user_msg = dict(new_hist[-1])
    user_msg["meta"] = {**(user_msg.get("meta") or {}), "turn_id": turn_id}
    msg_item = _ui._serialize_message_for_turn(user_msg)
    from agent.utils.agent_utils import save_turn_messages
    save_turn_messages(session_id, turn_id, [msg_item] if msg_item else [])

    ids = _ensure_session_in_list(session_id, ui.list_session_ids_ordered())
    return {
        "sessionId": session_id,
        "sessions": ids,
        "messages": _serialize_messages(new_hist, session_id),
    }


def _chat_complete_core(session_id: str, output_type: str = "text", agent_mode: str = "agent") -> dict[str, Any]:
    """对 turn 文件中末条 user 执行视觉解析与 ``chain_with_history.invoke``。

    输入:
        session_id: 会话 ID。
        output_type: text | voice。
        agent_mode: agent | plan。

    输出:
        {"sessionId": str, "sessions": list[str], "messages": list[dict]}

    系统定位:
        /api/chat/complete 与 /api/chat（兼容模式）的后端核心逻辑。
        图在后台线程（turn_runner）中执行，本函数阻塞等待其完成；
        期间人机交互浮窗与工具调用列表由 live snapshot 通道推送。

    可扩展性:
        可支持异步执行、流式输出。
    """
    hist = ui.load_ui_messages_for_session(session_id)
    if not hist or hist[-1].get("role") != "user":
        raise HTTPException(
            status_code=400,
            detail="没有待完成的用户消息，请先调用 /api/chat/start",
        )
    normalized_mode = agent_mode if agent_mode in {"agent", "plan"} else "agent"
    turn_runner.start_turn(session_id, [*hist], normalized_mode)
    result = turn_runner.wait_turn(session_id)
    if result is None or "error" in result:
        raise HTTPException(status_code=500, detail="Agent 执行失败")
    new_hist = result["chat_history"]
    ui.prune_memory_session_ids()
    ids = _ensure_session_in_list(session_id, ui.list_session_ids_ordered())
    return {
        "sessionId": session_id,
        "sessions": ids,
        "messages": _serialize_messages(new_hist, session_id),
    }


@app.post("/api/chat/start")
async def api_chat_start(
    session_id: str = Form(...),
    text: str = Form(""),
    images: list[UploadFile] | None = File(None),
    files: list[UploadFile] | None = File(None),
    voice_input: UploadFile | None = File(None),
) -> dict[str, Any]:
    """追加用户消息并持久化到 turn 文件（含附件落盘）；不调用模型。

    前端应先调用本接口再调用 ``/api/chat/complete``。

    输入:
        session_id: 表单字段，会话 ID。
        text: 用户输入的文本。
        images: 上传的图片文件列表。
        files: 上传的普通文件列表（含音频文件附件，作为普通文件展示）。
        voice_input: 麦克风录音的语音输入文件；存在时阻塞执行 ASR 转文本，
            转出的文字并入 ``text``，音频本身不持久化。

    输出:
        {"sessionId": str, "sessions": list[str], "messages": list[dict]}

    系统定位:
        两步式对话的第一步，用于提前上传文件并显示用户消息。
        语音输入在此处阻塞转文字，前端期间显示"语音转文字中…"占位。
    """
    if not session_id:
        raise HTTPException(status_code=400, detail="缺少 session_id")

    # 语音输入：阻塞 ASR 转文本，并入 text
    voice_path, voice_tmp = _save_voice_input_to_temp(voice_input)
    try:
        final_text = (text or "").strip()
        if voice_path:
            asr_text = _transcribe_voice_input(voice_path)
            if asr_text:
                final_text = (final_text + "\n" + asr_text).strip() if final_text else asr_text
    finally:
        shutil.rmtree(voice_tmp, ignore_errors=True)

    query, tmp_root = _materialize_uploads_to_temp(images, files, final_text)
    try:
        return _chat_start_core(session_id, query)
    finally:
        shutil.rmtree(tmp_root, ignore_errors=True)


@app.post("/api/chat/complete")
def api_chat_complete(
    session_id: str = Form(...),
    output_type: str = Form("text"),
    agent_mode: str = Form("agent"),
) -> dict[str, Any]:
    """对当前会话最近一次 ``/api/chat/start`` 提交执行模型推理并返回完整消息列表。

    输入:
        session_id: 表单字段，会话 ID。
        output_type: text | voice。
        agent_mode: agent | plan。

    输出:
        {"sessionId": str, "sessions": list[str], "messages": list[dict]}
    """
    if not session_id:
        raise HTTPException(status_code=400, detail="缺少 session_id")
    return _chat_complete_core(session_id, output_type=output_type, agent_mode=agent_mode)


def _stream_chat_complete_core(session_id: str, output_type: str = "text", agent_mode: str = "agent"):
    """SSE 流式推理：后台线程执行 Agent，完成后对最终回复做流式 TTS。"""
    import time as _time

    # 启动后台线程执行 Agent（阻塞等待；期间 live snapshot 推送工具调用与人机交互浮窗）
    hist = ui.load_ui_messages_for_session(session_id)
    if not hist or hist[-1].get("role") != "user":
        yield f"data: {json.dumps({'type': 'error', 'message': '没有待完成的用户消息'}, ensure_ascii=False)}\n\n"
        return

    normalized_mode = agent_mode if agent_mode in {"agent", "plan"} else "agent"
    turn_runner.start_turn(session_id, [*hist], normalized_mode)
    result = turn_runner.wait_turn(session_id)
    if result is None or "error" in result:
        yield f"data: {json.dumps({'type': 'error', 'message': 'Agent 执行失败'}, ensure_ascii=False)}\n\n"
        return
    new_hist = result["chat_history"]
    ui.prune_memory_session_ids()
    ids = _ensure_session_in_list(session_id, ui.list_session_ids_ordered())

    # 提取最终 assistant 消息文本
    final_text = ""
    messages = _serialize_messages(new_hist, session_id)
    for m in reversed(messages):
        if m.get("role") == "assistant":
            final_text = m.get("content", "")
            if isinstance(final_text, list):
                final_text = "".join(p.get("text", "") for p in final_text if isinstance(p, dict) and p.get("type") == "text")
            break

    # 先发送文本消息（含完整的 messages、sessions）
    yield f"data: {json.dumps({'type': 'text', 'text': str(final_text), 'messages': messages, 'sessionId': session_id, 'sessions': ids}, ensure_ascii=False)}\n\n"

    # 语音模式：流式 TTS
    if output_type == "voice" and final_text.strip():
        request_id = f"tts-{session_id}-{int(_time.time() * 1000)}"
        tts_url = "http://localhost:8902"
        try:
            from agent.utils.env_utils import STREAMING_TTS_URL
            tts_url = STREAMING_TTS_URL or tts_url

            # 清理文本 + 按句号换行（帮助 VoxCPM 分句）
            tts_text = _clean_tts_text(final_text)

            import requests as _req
            resp = _req.post(
                f"{tts_url}/stream_tts",
                json={"text": tts_text, "request_id": request_id},
                stream=True,
                timeout=300,
            )
            resp.raise_for_status()

            for line in resp.iter_lines(decode_unicode=True):
                if not line or not line.startswith("data: "):
                    continue
                payload_str = line[len("data: "):]
                yield f"data: {payload_str}\n\n"

        except Exception as e:
            yield f"data: {json.dumps({'type': 'audio.error', 'message': str(e)}, ensure_ascii=False)}\n\n"

    yield f"data: {json.dumps({'type': 'done'}, ensure_ascii=False)}\n\n"


@app.post("/api/chat/stream")
def api_chat_stream(
    session_id: str = Form(...),
    output_type: str = Form("text"),
    agent_mode: str = Form("agent"),
):
    """SSE 流式推理 + 流式 TTS。

    输入:
        session_id: 表单字段，会话 ID。
        output_type: text | voice。
        agent_mode: agent | plan。

    输出:
        text/event-stream，事件类型：text / audio.chunk / audio.done / done。
    """
    if not session_id:
        raise HTTPException(status_code=400, detail="缺少 session_id")
    return StreamingResponse(
        _stream_chat_complete_core(session_id, output_type=output_type, agent_mode=agent_mode),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@app.post("/api/chat/abort")
def api_chat_abort(session_id: str = Form(...)) -> dict[str, Any]:
    """中止当前正在执行的 Agent 任务，保留已完成记录，以暂停消息结束本轮。

    输入:
        session_id: 表单字段，会话 ID。

    输出:
        {"sessionId": str, "sessions": list[str], "messages": list[dict]}
    """
    if not session_id:
        raise HTTPException(status_code=400, detail="缺少 session_id")
    try:
        from agent.core.runtime import set_abort_flag
        set_abort_flag(session_id)
    except Exception:
        pass
    # 应答等待中的人工请求：阻塞工具醒来后检测到 abort 标志，抛 ABORTED_BY_USER
    # 由 runtime 统一按暂停处理（先置标志再应答，保证暂停优先于答案）
    try:
        from agent.core.human_request import resolve_all_for_session
        resolve_all_for_session(session_id, decision="reject", instruction="已被用户手动暂停")
    except Exception:
        pass
    # 不 rollback，保留当前轮已完成的工具调用记录，追加暂停消息结束本轮
    hist = ui.load_ui_messages_for_session(session_id)
    paused = list(hist or [])
    # 避免重复追加（execute_agent 可能已追加过）
    last_msg = paused[-1] if paused else None
    if not (last_msg and last_msg.get("role") == "assistant" and last_msg.get("content") == "已被用户手动暂停"):
        paused.append({"role": "assistant", "content": "已被用户手动暂停", "meta": {"paused": True}})
    ui.sync_turn_from_chat_history(session_id, paused, "")
    ids = _ensure_session_in_list(session_id, ui.list_session_ids_ordered())
    return {
        "sessionId": session_id,
        "sessions": ids,
        "messages": _serialize_messages(paused, session_id),
    }


@app.post("/api/chat/rollback-to-message")
def api_chat_rollback_to_message(
    session_id: str = Form(...),
    rollback_index: int = Form(...),
) -> dict[str, Any]:
    """将聊天历史回撤到指定消息之前，清理相关数据。

    输入:
        session_id: 会话 ID。
        rollback_index: 要回退到的消息在 chat_history 中的索引（该消息会被移除并返回给前端放回输入区）。

    输出:
        {"sessionId": str, "sessions": list[str], "messages": list[dict],
         "rollbackMessage": dict | None}
    """
    if not session_id:
        raise HTTPException(status_code=400, detail="缺少 session_id")

    # 1. 加载历史，回滚到指定索引（不含该消息）
    hist = ui.load_ui_messages_for_session(session_id)
    if rollback_index < 0 or rollback_index >= len(hist):
        raise HTTPException(status_code=400, detail="无效的回撤索引")

    rollback_msg = hist[rollback_index] if rollback_index < len(hist) else None
    rolled_back = list(hist[:rollback_index])  # 不含回撤点消息

    # 2. 收集需要保留的 turn_id（回滚后仍存在的消息对应的 turn）
    keep_turn_ids = set()
    for msg in rolled_back:
        tid = msg.get("meta", {}).get("turn_id") if isinstance(msg.get("meta"), dict) else None
        if tid:
            keep_turn_ids.add(tid)
    # 清理 tool_calling（保留未回滚部分的工具调用记录）、删除多余的 turn 目录
    _clear_tool_calling(session_id, keep_turn_ids=keep_turn_ids)
    from agent.utils.agent_utils import delete_turns_after
    # 找到最后一个要保留的 turn_id
    keep_turn_id = "00000000_000000"
    if rolled_back:
        for msg in reversed(rolled_back):
            tid = msg.get("meta", {}).get("turn_id") if isinstance(msg.get("meta"), dict) else None
            if tid:
                keep_turn_id = tid
                break
    delete_turns_after(session_id, keep_turn_id)

    # 3. 同步 turn 文件（写最新的 turn 状态）

    ids = _ensure_session_in_list(session_id, ui.list_session_ids_ordered())
    return {
        "sessionId": session_id,
        "sessions": ids,
        "messages": _serialize_messages(rolled_back, session_id),
        "rollbackMessage": _serialize_messages([rollback_msg], session_id)[0] if rollback_msg else None,
    }


@app.post("/api/human-action")
def api_human_action(
    session_id: str = Form(...),
    approval_id: str = Form(...),
    decision: str = Form(...),
    instruction: str = Form(""),
) -> dict[str, Any]:
    """应答一个等待中的人工请求（人机交互/工具确认）。

    输入:
        session_id: 会话 ID。
        approval_id: 等待中请求的 ID（由 live snapshot 的 human_requests 提供）。
        decision: approve | reject | edit | skip（skip 表示跳过工具调用但继续图）。
        instruction: 用户输入的补充文本（用于 edit 或 selection）。

    输出:
        {"sessionId": str, "status": "answered"}

    系统定位:
        前端浮窗提交决策时调用的回调接口。答案写入等待注册表并唤醒阻塞中的
        工具，图在后台线程自然继续；最终结果由 /api/chat/complete 或
        /api/chat/stream 返回，本接口不做任何图执行。
    """
    if not session_id:
        raise HTTPException(status_code=400, detail="缺少 session_id")
    if decision not in {"approve", "reject", "edit", "skip"}:
        raise HTTPException(status_code=400, detail="不支持的人工处理结果")
    from agent.core.human_request import answer_human_request
    ok = answer_human_request(session_id, approval_id, decision, instruction)
    if not ok:
        raise HTTPException(status_code=404, detail="该待确认项已处理或已过期")
    return {"sessionId": session_id, "status": "answered"}


# ---------- 配置管理 API ----------
@app.get("/api/config/agents")
def api_get_agent_configs() -> dict[str, Any]:
    """获取所有 agent 配置列表。

    输出: {"agents": list[dict], "config_path": str}
    """
    from agent.core.config_manager import load_models
    configs = load_agent_configs()
    models = {m["id"]: m for m in load_models()}
    return {
        "agents": [{"name": c.name, "description": c.description, "system_prompt": c.system_prompt, "max_iterations": c.max_iterations,
                     "llm_model_id": c.llm_model_id,
                     "tools": c.tools,
                     "enabled": c.enabled,
                     "role": c.role,
                     "trajectory_rounds": c.trajectory_rounds} for c in configs],
        "models": list(models.values()),
        "config_path": str(AGENT_CONFIG_PATH),
    }


@app.post("/api/config/agents")
def api_save_agent_configs(data: dict[str, Any]) -> dict[str, Any]:
    """保存 agent 配置列表。

    输入: {"agents": list[dict]}
    输出: {"success": bool, "config_path": str}
    """
    agents = data.get("agents") if isinstance(data, dict) else None
    if not isinstance(agents, list):
        raise HTTPException(status_code=400, detail="缺少 agents 字段或格式不正确")
    success = save_agent_configs(agents)
    return {"success": success, "config_path": str(AGENT_CONFIG_PATH)}


@app.get("/api/config/tools")
def api_get_tool_configs() -> dict[str, Any]:
    """获取所有工具配置列表。

    输出: {"tools": list[dict], "registered_tool_names": list[str],
            "tool_parameters": dict, "forced_permissions": dict, "config_path": str}
    """
    from agent.core.config_manager import _get_tool_parameters, get_tool_forced_permissions
    configs = load_tool_configs()
    return {
        "tools": [{"name": c.name, "display_name": c.display_name, "description": c.description,
                    "permission": c.permission, "auto_execute_rules": c.auto_execute_rules, "enabled": c.enabled,
                    "timeout": c.timeout}
                  for c in configs],
        "registered_tool_names": get_registered_tool_names(),
        "tool_parameters": _get_tool_parameters(),
        "forced_permissions": get_tool_forced_permissions(),
        "config_path": str(TOOL_CONFIG_PATH),
    }


@app.post("/api/config/tools")
def api_save_tool_configs(data: dict[str, Any]) -> dict[str, Any]:
    """保存工具配置列表。

    输入: {"tools": list[dict]}
    输出: {"success": bool, "config_path": str}
    """
    tools = data.get("tools") if isinstance(data, dict) else None
    if not isinstance(tools, list):
        raise HTTPException(status_code=400, detail="缺少 tools 字段或格式不正确")
    success = save_tool_configs(tools)
    # 保存后清除 tool_policy 的规则缓存，使新配置生效
    try:
        from agent.core.tool_policy import reload_auto_execute_rules
        reload_auto_execute_rules()
    except Exception:
        pass
    return {"success": success, "config_path": str(TOOL_CONFIG_PATH)}


# ---------- 模型管理 API ----------
@app.get("/api/config/models")
def api_get_models() -> dict[str, Any]:
    """获取所有 LLM 模型配置。

    输出: {"models": list[dict]}
    """
    return {"models": load_models()}


@app.post("/api/config/models")
def api_save_models(data: dict[str, Any]) -> dict[str, Any]:
    """保存 LLM 模型配置列表。

    输入: {"models": list[dict]}
    输出: {"success": bool}
    """
    models = data.get("models") if isinstance(data, dict) else None
    if not isinstance(models, list):
        raise HTTPException(status_code=400, detail="缺少 models 字段或格式不正确")
    success = save_models(models)
    return {"success": success}


# ---------- 工作空间 API ----------

@app.get("/api/workspace")
def api_get_workspace() -> dict[str, Any]:
    """获取工作空间配置。

    输出: {"current": str, "list": list[str], "default": str, "snapshots_base": str,
           "access_mode": str}
        current: 当前选中的工作空间路径（空字符串表示使用默认值）
        list: 已保存的工作空间路径列表
        default: env_config 中的默认工作空间路径
        snapshots_base: 快照存储基础目录
        access_mode: 工作空间访问模式（restricted | approval | full）
    """
    from agent.utils.env_utils import get_workspace_dir
    from agent.memory.system_prompt import _current_workspace_dir
    from agent.utils.agent_utils import SESSIONS_ROOT
    from agent.core.workspace_policy import get_access_mode
    import os
    cfg = _read_ws_config()
    current = _current_workspace_dir or cfg.get("current", "")
    ws_list = cfg.get("list", [])

    # 清理列表中已不存在的文件夹
    valid = [p for p in ws_list if os.path.isdir(p)]
    if len(valid) != len(ws_list):
        cfg["list"] = valid
        _write_ws_config(cfg)

    return {
        "current": current,
        "list": valid,
        "default": get_workspace_dir(),
        "snapshots_base": os.path.join(SESSIONS_ROOT, ".minor_snapshots").replace("\\", "/"),
        "access_mode": get_access_mode(),
    }


@app.post("/api/workspace/access-mode")
def api_set_access_mode(data: dict[str, Any]) -> dict[str, Any]:
    """设置工作空间访问模式。

    输入: {"mode": "restricted"|"approval"|"full"}
    输出: {"success": bool, "mode": str}
    """
    from agent.core.workspace_policy import _ACCESS_MODES, set_access_mode
    mode = data.get("mode", "") if isinstance(data, dict) else ""
    if mode not in _ACCESS_MODES:
        raise HTTPException(status_code=400, detail="无效的访问模式（restricted | approval | full）")
    ok = set_access_mode(mode)
    if not ok:
        raise HTTPException(status_code=500, detail="保存访问模式失败")
    return {"success": True, "mode": mode}


@app.post("/api/workspace/select")
def api_select_workspace(data: dict[str, Any]) -> dict[str, Any]:
    """选择工作空间。

    输入: {"path": str}  - 工作空间绝对路径，空字符串表示使用默认值
    输出: {"success": bool, "current": str}
    """
    from agent.memory.system_prompt import set_current_workspace
    path = data.get("path", "") if isinstance(data, dict) else ""
    cfg = _read_ws_config()
    cfg["current"] = path
    _write_ws_config(cfg)
    set_current_workspace(path if path else None)
    # 同步更新 env_utils.WORKSPACE_DIR，确保 image_gen / terminal_execute 使用正确路径
    try:
        import agent.utils.env_utils as env_utils
        env_utils.WORKSPACE_DIR = path if path else ""
    except Exception:
        pass
    # 重建 Agent 运行时以应用新的系统提示词
    try:
        from agent.agents.agent_runtime import reload_all_agent_runtimes
        reload_all_agent_runtimes()
    except Exception:
        pass
    return {"success": True, "current": path}


@app.post("/api/workspace/add")
def api_add_workspace(data: dict[str, Any]) -> dict[str, Any]:
    """添加工作空间到列表。

    输入: {"path": str}  - 工作空间绝对路径
    输出: {"success": bool, "list": list[str]}
    """
    path = data.get("path", "") if isinstance(data, dict) else ""
    if not path or not os.path.isdir(path):
        raise HTTPException(status_code=400, detail="无效的工作空间路径")
    cfg = _read_ws_config()
    ws_list = cfg.get("list", [])
    if path not in ws_list:
        ws_list.insert(0, path)
        if len(ws_list) > 20:
            ws_list = ws_list[:20]
        cfg["list"] = ws_list
        _write_ws_config(cfg)
    return {"success": True, "list": ws_list}


@app.post("/api/workspace/remove")
def api_remove_workspace(data: dict[str, Any]) -> dict[str, Any]:
    """从列表中移除工作空间。

    输入: {"path": str}  - 要移除的工作空间路径
    输出: {"success": bool, "list": list[str], "current": str}
    """
    from agent.memory.system_prompt import set_current_workspace, _current_workspace_dir
    path = data.get("path", "") if isinstance(data, dict) else ""
    cfg = _read_ws_config()
    ws_list = cfg.get("list", [])
    ws_list = [p for p in ws_list if p != path]
    cfg["list"] = ws_list
    # 如果移除的是当前选中的，恢复默认
    if _current_workspace_dir == path or cfg.get("current") == path:
        cfg["current"] = ""
        set_current_workspace(None)
        # 同步 env_utils.WORKSPACE_DIR 恢复为 env_config 中的默认值
        try:
            import agent.utils.env_utils as env_utils
            from agent.utils.env_utils import get_workspace_dir
            env_utils.WORKSPACE_DIR = get_workspace_dir()
        except Exception:
            pass
    _write_ws_config(cfg)
    return {"success": True, "list": ws_list, "current": cfg.get("current", "")}


@app.get("/api/config/env")
def api_get_env_config() -> dict[str, Any]:
    """获取环境变量配置。

    输出: {"env": dict, "models": list[dict], "config_path": str}
    """
    config = load_env_config()
    models = config.get("models", [])
    env_only = {k: v for k, v in config.items() if k != "models"}
    return {"env": env_only, "models": models, "config_path": str(ENV_CONFIG_PATH)}


@app.post("/api/config/env")
def api_save_env_config(data: dict[str, Any]) -> dict[str, Any]:
    """保存环境变量配置。

    输入: {"env": dict}
    输出: {"success": bool, "config_path": str}
    """
    env = data.get("env") if isinstance(data, dict) else None
    if not isinstance(env, dict):
        raise HTTPException(status_code=400, detail="缺少 env 字段或格式不正确")
    success = save_env_config(env)
    return {"success": success, "config_path": str(ENV_CONFIG_PATH)}


# ---------- GUI 显示器配置 API ----------
@app.get("/api/config/gui")
def api_get_gui_config() -> dict[str, Any]:
    """获取 GUI 显示器配置 + 实时显示器列表 + 可用模型列表。

    输出: {"monitors": list[dict], "selected_name": str, "gui_model_id": str, "models": list[dict], "config_path": str}
    """
    config = load_gui_config()
    live_monitors = _get_live_monitors()
    models = load_models()
    return {
        "monitors": live_monitors,
        "selected_name": config.get("gui_monitor_name", ""),
        "gui_model_id": config.get("gui_model_id", ""),
        "models": models,
        "config_path": str(GUI_CONFIG_PATH),
    }


@app.post("/api/config/gui")
def api_save_gui_config(data: dict[str, Any]) -> dict[str, Any]:
    """保存 GUI 显示器选择和 GUI 模型。

    输入: {"gui_monitor_name": str, "gui_model_id": str}
    输出: {"success": bool, "config_path": str}
    """
    gui_monitor_name = data.get("gui_monitor_name") if isinstance(data, dict) else None
    gui_model_id = data.get("gui_model_id") if isinstance(data, dict) else None
    if not isinstance(gui_monitor_name, str) or not gui_monitor_name.strip():
        raise HTTPException(status_code=400, detail="缺少 gui_monitor_name 字段或格式不正确")
    config = load_gui_config()
    config["gui_monitor_name"] = gui_monitor_name.strip()
    if isinstance(gui_model_id, str):
        config["gui_model_id"] = gui_model_id.strip()
    success = save_gui_config(config)
    return {"success": success, "config_path": str(GUI_CONFIG_PATH)}


# 先注册 API，再挂静态文件，避免 ``/`` 抢路由
@app.get("/api/health")
def health() -> dict[str, str]:
    """健康检查接口。

    输入: 无。

    输出: {"status": "ok"}
    """
    return {"status": "ok"}


@app.get("/api/services/status")
def api_services_status() -> dict[str, Any]:
    """探测 ASR 与 TTS 服务的连通性（供前端麦克风/语音模式切换前检查）。

    探测方式: 对服务根路径发起 GET，任何 HTTP 响应（含 404）都视为服务在线；
    连接失败/超时视为服务不可用。

    输出: {"asr": {"ok": bool, "url": str}, "tts": {"ok": bool, "url": str}}
    """
    import requests as _req
    from agent.utils.env_utils import ASR_BASE_URL, STREAMING_TTS_URL

    def _probe(url: str) -> bool:
        if not url or not str(url).strip():
            return False
        try:
            _req.get(str(url).rstrip("/") + "/", timeout=3)
            return True
        except Exception:
            return False

    return {
        "asr": {"ok": _probe(ASR_BASE_URL), "url": ASR_BASE_URL or ""},
        "tts": {"ok": _probe(STREAMING_TTS_URL), "url": STREAMING_TTS_URL or ""},
    }


@app.get("/api/pick-folder")
def api_pick_folder() -> dict[str, Any]:
    """打开原生文件夹选择对话框，返回用户选择的文件夹绝对路径。

    输入: 无。

    输出: {"path": str} 或 {"path": null}
    """
    import subprocess

    # 使用 PowerShell + Shell.Application COM 对象调出文件夹选择器
    ps_script = """
$ErrorActionPreference = 'Stop'
$shell = New-Object -ComObject Shell.Application
$folder = $shell.BrowseForFolder(0, '选择文件夹', 0, 0)
if ($folder) { Write-Output $folder.Self.Path }
"""
    try:
        proc = subprocess.run(
            ["powershell", "-NoProfile", "-NonInteractive", "-Command", ps_script],
            capture_output=True,
            text=True,
            timeout=120,
            encoding="gbk",
            errors="replace",
        )
        selected = proc.stdout.strip()
        if selected and os.path.isdir(selected):
            return {"path": selected}
    except Exception:
        pass
    return {"path": None}


@app.post("/api/config/reload")
def api_reload_configs() -> dict[str, Any]:
    """热重载所有配置并重建 Agent 运行时，无需重启服务。

    输入: 无（或 {}）。

    输出: {"success": bool, "agent": dict, "detail": str}
    """
    try:
        # 1. 重新读取 env_config.json 并写入 os.environ
        from agent.utils.env_utils import _load_config
        _load_config()

        # 2. 重建 LLM 实例
        from agent.core.llm import reload_llm
        reload_llm()

        # 3. 清除并重新加载 tool_policy 缓存
        from agent.core.tool_policy import reload_auto_execute_rules
        reload_auto_execute_rules()

        # 4. 清除工具超时缓存，使新的 timeout 配置生效
        from agent.core.config_manager import invalidate_tool_timeout_cache
        invalidate_tool_timeout_cache()

        # 5. 重建所有 Agent 运行时（从 agent_config.json 动态加载）
        from agent.agents.agent_runtime import reload_all_agent_runtimes
        agent_info = reload_all_agent_runtimes()

        return {"success": True, "agent": agent_info, "detail": "配置已热重载"}
    except Exception as e:
        return {"success": False, "agent": {}, "detail": str(e)}


# ---------- 压缩设置 API ----------
@app.get("/api/config/compress-settings")
def api_get_compress_settings() -> dict[str, Any]:
    """获取压缩相关设置。

    输出: {"context_window": int, "compress_rate": float}
    """
    from agent.utils.env_utils import LLM_CONTEXT_WINDOW, COMPRESS_RATE
    return {
        "context_window": LLM_CONTEXT_WINDOW,
        "compress_rate": COMPRESS_RATE,
    }


@app.post("/api/config/compress-settings")
def api_save_compress_settings(data: dict[str, Any]) -> dict[str, Any]:
    """更新压缩参数（COMPRESS_RATE）。

    输入: {"compress_rate": float}
    输出: {"success": bool}
    """
    config = load_env_config()
    updated = False
    if "compress_rate" in data:
        config["COMPRESS_RATE"] = str(data["compress_rate"])
        updated = True
    if updated:
        save_env_config(config)
        from agent.utils.env_utils import _load_config
        _load_config()
    return {"success": updated}


# ---------- 深度思考档位 API ----------
@app.post("/api/config/thinking-level")
def api_set_thinking_level(data: dict[str, Any]) -> dict[str, Any]:
    """设置深度思考档位（low | high | xhigh | max | ultra），重建 Agent 运行时。

    档位同时决定运行模式与思考开关：
        low = agent + 不思考；high = plan + 不思考；
        xhigh = agent + 思考；max = plan + 思考；
        ultra 为未来 react→审批图预留占位（暂按 agent + 思考）。

    输入: {"level": str}
    输出: {"success": bool, "level": str}
    """
    from agent.utils.env_utils import THINKING_LEVELS, _load_config
    level = data.get("level", "") if isinstance(data, dict) else ""
    if level not in THINKING_LEVELS:
        raise HTTPException(status_code=400, detail="无效的思考档位（low | high | xhigh | max | ultra）")
    config = load_env_config()
    config["THINKING_LEVEL"] = level
    if not save_env_config(config):
        raise HTTPException(status_code=500, detail="保存思考档位失败")
    _load_config()
    # 重建 Agent 运行时使新的 extra_body 生效
    try:
        from agent.agents.agent_runtime import reload_all_agent_runtimes
        reload_all_agent_runtimes()
    except Exception:
        pass
    return {"success": True, "level": level}


# ---------- 会话 Token 用量 API ----------
@app.get("/api/chat/tokens/{session_id}")
def api_get_session_tokens(session_id: str) -> dict[str, Any]:
    """获取当前会话的 token 用量（来自 LLM response usage_metadata.total_tokens）。

    输出: {"tokens": int, "context_window": int, "compress_threshold": int,
           "compress_rate": float}
    """
    from agent.utils.env_utils import LLM_CONTEXT_WINDOW, COMPRESS_RATE
    from agent.history.tool_call_recorder import get_session_tokens, get_session_token_breakdown

    # 直接读取 LLM 每次调用后存储的真实 total_tokens
    tokens = get_session_tokens(session_id)
    compress_threshold = int(LLM_CONTEXT_WINDOW * COMPRESS_RATE)

    return {
        "tokens": tokens,
        "context_window": LLM_CONTEXT_WINDOW,
        "compress_threshold": compress_threshold,
        "compress_rate": COMPRESS_RATE,
        # 三类估算占比（消息/工具/系统提示词），供前端环形指示展示
        "breakdown": get_session_token_breakdown(session_id),
    }


# ---------- 全局 Token 用量统计 API ----------
@app.get("/api/stats/usage")
def api_get_usage_stats(days: int = 182) -> dict[str, Any]:
    """获取全局 token 用量统计（跨会话累加，欢迎区活跃度矩阵 / 柱状图数据源）。

    输出:
        {"total": int, "by_model": {model: int},
         "days": [{"date": str, "total": int, "by_model": {model: int}}],
         "hours": {"YYYY-MM-DD": {"HH": {model: int}}}}
    """
    from agent.core.usage_stats import get_usage_summary
    return get_usage_summary(days=max(1, min(days, 366)))


# ---------- 工具调用历史 API ----------
@app.get("/api/tool-calls/{session_id}/live")
def api_get_live_tool_calls(session_id: str) -> dict[str, Any]:
    """获取当前进行中的工具调用、反思内容与等待中的人工请求（前端执行期间实时轮询）。"""
    try:
        from agent.history.tool_call_recorder import get_live_tool_calls, get_live_reflections, get_turn_started_at
        from agent.core.human_request import get_human_requests_for_session
        return {
            "tool_calls": get_live_tool_calls(session_id),
            "reflections": get_live_reflections(session_id),
            "started_at": get_turn_started_at(session_id),
            "human_requests": get_human_requests_for_session(session_id),
        }
    except Exception as e:
        return {"tool_calls": [], "reflections": [], "human_requests": [], "error": str(e)}


@app.get("/api/tool-calls/{session_id}/stream")
async def api_stream_live_tool_calls(session_id: str, request: Request):
    """SSE 流式推送实时工具调用更新（触发式，替代轮询）。

    每当后端 record_tool_call_live / record_tool_result_live / record_reflection_live
    或人工请求注册/应答被调用时，通过 pub/sub queue 立即推送快照给前端。
    """
    from agent.history.tool_call_recorder import (
        subscribe_live, unsubscribe_live,
        get_live_tool_calls, get_live_reflections, get_turn_started_at,
    )
    from agent.core.human_request import get_human_requests_for_session

    q = subscribe_live(session_id)

    async def event_stream():
        try:
            # 发送初始快照（连接时可能已有数据）
            initial = {
                "tool_calls": get_live_tool_calls(session_id),
                "reflections": get_live_reflections(session_id),
                "started_at": get_turn_started_at(session_id),
                "human_requests": get_human_requests_for_session(session_id),
            }
            yield f"event: update\ndata: {json.dumps(initial, ensure_ascii=False)}\n\n"

            while True:
                if await request.is_disconnected():
                    break
                try:
                    # 在线程中等待队列，超时后发送心跳
                    snapshot = await asyncio.to_thread(q.get, True, 3.0)
                    yield f"event: update\ndata: {json.dumps(snapshot, ensure_ascii=False)}\n\n"
                except _queue_module.Empty:
                    # 心跳保活
                    yield f"event: ping\ndata: \n\n"
        finally:
            unsubscribe_live(session_id, q)

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


@app.get("/api/tool-image")
def api_get_tool_image(path: str, session: str = "") -> FileResponse:
    """提供工具调用历史中的图片文件。

    输入:
        path: 图片文件路径。
        session: 兼容参数。

    输出:
        图片文件响应。
    """
    try:
        img_path = Path(path).resolve()
    except Exception:
        raise HTTPException(status_code=400, detail="无效的图片路径")
    if not img_path.is_file():
        raise HTTPException(status_code=404, detail="工具调用图片不存在")
    mime, _ = mimetypes.guess_type(str(img_path))
    return FileResponse(img_path, media_type=mime or "image/png")


@app.get("/api/tool-file")
def api_get_tool_file(path: str, session: str = "") -> FileResponse:
    """提供工具调用历史中的非图片文件（下载或播放）。

    输入:
        path: 文件路径（可为 turn 目录下的落盘文件或原始文件路径）。
        session: 兼容参数。

    输出:
        带 Content-Disposition 的文件响应，浏览器可下载或播放。
    """
    try:
        file_path = Path(path).resolve()
    except Exception:
        raise HTTPException(status_code=400, detail="无效的文件路径")
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="文件不存在")
    mime, _ = mimetypes.guess_type(str(file_path))
    media_type = mime or "application/octet-stream"
    # 图片类 inline 展示，其余触发下载
    if media_type.startswith("image/") or media_type.startswith("audio/") or media_type.startswith("video/"):
        return FileResponse(file_path, media_type=media_type)
    return FileResponse(
        file_path,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{file_path.name}"'},
    )


# ---------- Skill 管理 API ----------
@app.get("/api/skills")
def api_list_skills() -> dict[str, Any]:
    """列出所有 skill 的基本信息。

    输出: {"skills": list[dict]}  每项含 name, description, version, author, tags, has_attachments, attachment_files
    """
    from agent.skills import list_skills
    return {"skills": list_skills()}


@app.get("/api/skills/{name}")
def api_get_skill(name: str) -> dict[str, Any]:
    """获取单个 skill 的完整信息（含 skill.md 内容）。

    输出: {"skill": dict}  含 name, description, version, author, tags, content, attachment_files
    """
    from agent.skills import get_skill
    skill = get_skill(name)
    if skill is None:
        raise HTTPException(status_code=404, detail=f"Skill '{name}' 不存在")
    return {"skill": skill}


@app.post("/api/skills")
async def api_create_or_update_skill(
    name: str = Form(...),
    description: str = Form(""),
    content: str = Form(""),
    tags: str = Form(""),
    enable: str = Form("true"),
) -> dict[str, Any]:
    """创建或更新一个 skill（不保留 version/author）。

    输入:
        name: skill 名称（必填）。
        description: 简述。
        content: skill.md 的 markdown 内容。
        tags: 逗号分隔的标签字符串（可选，便于分级筛选）。
        enable: 是否启用，默认 "true"。

    输出: {"success": bool, "skill": dict}
    """
    if not name or not name.strip():
        raise HTTPException(status_code=400, detail="name 不能为空")
    tag_list = [t.strip() for t in tags.split(",") if t.strip()] if tags else []
    enable_bool = enable.lower() not in ("false", "0", "no", "")
    from agent.skills import create_or_update_skill
    ok = create_or_update_skill(
        name=name.strip(),
        description=description,
        content=content,
        tags=tag_list,
        enable=enable_bool,
    )
    if not ok:
        raise HTTPException(status_code=500, detail="保存 skill 失败")
    from agent.skills import get_skill
    skill = get_skill(name.strip())
    return {"success": True, "skill": skill}


# ---------- Skill 导入（zip：先解析展示，保存时才落盘） ----------
@app.post("/api/skills/parse")
async def api_parse_skill_zip(file: UploadFile = File(...)) -> dict[str, Any]:
    """解析 skill zip 包（不落盘），返回 name/description/content 与附件列表。

    输入: multipart file（zip）
    输出: {"found": bool, "skill": {name, description, content, attachments} | None,
           "error": str | None}
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="文件名为空")
    zip_bytes = await file.read()
    from agent.skills import parse_skill_zip
    parsed = parse_skill_zip(zip_bytes)
    if parsed is None:
        return {"found": False, "skill": None,
                "error": "未在 zip 的一级子目录下找到 skill.md/SKILL.md"}
    return {"found": True, "skill": parsed, "error": None}


@app.post("/api/skills/import")
async def api_import_skill_zip(
    file: UploadFile = File(...),
    name: str = Form(""),
    description: str = Form(""),
    content: str = Form(""),
    tags: str = Form(""),
    enable: str = Form("true"),
) -> dict[str, Any]:
    """导入 skill zip 包到 skills 目录（skill.md/skill.json + 附件保持原相对路径）。

    输入: multipart file（zip）+ name/description/content/tags/enable 覆盖值
    输出: {"success": bool, "skill": dict | None}
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="文件名为空")
    zip_bytes = await file.read()
    from agent.skills import import_skill_zip, get_skill
    tag_list = [t.strip() for t in tags.split(",") if t.strip()] if tags else []
    enable_bool = enable.lower() not in ("false", "0", "no", "")
    ok = import_skill_zip(
        name=name.strip(),
        zip_bytes=zip_bytes,
        description=description,
        content=content,
        tags=tag_list,
        enable=enable_bool,
    )
    if not ok:
        raise HTTPException(status_code=500, detail="导入 skill 失败（zip 无效或未找到 skill.md）")
    return {"success": True, "skill": get_skill(name.strip())}


@app.delete("/api/skills/{name}")
def api_delete_skill(name: str) -> dict[str, Any]:
    """删除一个 skill 及其所有附件。

    输出: {"success": bool}
    """
    from agent.skills import delete_skill
    ok = delete_skill(name)
    return {"success": ok}


# ---------- Skill 附件管理 API ----------
@app.post("/api/skills/{name}/attachments")
async def api_add_skill_attachment(
    name: str,
    file: UploadFile = File(...),
) -> dict[str, Any]:
    """上传一个附件到指定 skill。

    输出: {"success": bool, "filename": str}
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="文件名为空")
    content = await file.read()
    from agent.skills import add_attachment
    ok = add_attachment(name, file.filename, content)
    if not ok:
        raise HTTPException(status_code=500, detail="上传附件失败（skill 不存在或路径非法）")
    return {"success": True, "filename": file.filename}


@app.delete("/api/skills/{name}/attachments/{filename:path}")
def api_delete_skill_attachment(name: str, filename: str) -> dict[str, Any]:
    """删除 skill 下的一个附件。

    输出: {"success": bool}
    """
    from agent.skills import delete_attachment
    ok = delete_attachment(name, filename)
    return {"success": ok}


def _clear_tool_calling(session_id: str, keep_turn_ids: set[str] | None = None) -> None:
    """清理会话关联的工具调用记录。

    输入:
        session_id: 会话 ID。
        keep_turn_ids: 需要保留的 turn_id 集合（回滚时保留未回滚部分的工具调用记录）。
                       为 None 时删除全部工具调用记录。

    系统定位:
        工具调用记录删除由统一存储处理后端无关（mysql 删表 / json 删 tool_*.json）；
        artifact 目录（以 turn_id 命名的目录，两个后端均落盘）在此清理。
    """
    try:
        from agent.history import session_storage
        session_storage.delete_tool_records(session_id, keep_turn_ids)
    except Exception:
        pass
    try:
        from pathlib import Path as _Path
        from agent.history.tool_call_recorder import session_tool_dir
        sd = _Path(session_tool_dir(session_id))
        if sd.is_dir():
            import shutil as _shutil
            # 只删除 artifact 目录（以 turn_id 命名，不含 turn_ 前缀的目录），
            # 不删 turn JSON（对话内容）和 session_meta
            for sub in sd.iterdir():
                if sub.is_dir():
                    if not sub.name.startswith("turn_"):
                        _shutil.rmtree(str(sub), ignore_errors=True)
    except Exception:
        pass


@app.get("/favicon.ico")
def favicon() -> FileResponse:
    """返回网站图标（favicon.ico/png/svg）。"""
    for name, mime in (
        ("favicon.ico", "image/x-icon"),
        ("favicon.png", "image/png"),
        ("favicon.svg", "image/svg+xml"),
    ):
        path = _IMAGE_DIR / name
        if path.is_file():
            return FileResponse(path, media_type=mime)
    raise HTTPException(status_code=404, detail="web/image 目录下缺少 favicon.ico / favicon.png / favicon.svg")


@app.get("/api/config/theme")
def get_theme_config() -> dict:
    """读取主题配置。"""
    import json
    config_path = _CONFIG_DIR / "theme_config.json"
    if config_path.is_file():
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"theme": "cyber", "bg_image": "", "font_family": "msjh", "font_size": "14px", "language": "zh"}


@app.post("/api/config/theme")
async def save_theme_config(request: Request):
    """保存主题配置。"""
    config_path = _CONFIG_DIR / "theme_config.json"
    _CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    data = await request.json()
    with open(config_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return {"ok": True}


@app.get("/api/bg-image")
def serve_local_bg_image(path: str = "") -> FileResponse:
    """提供本地任意路径的图片文件（用于自定义背景图）。

    输入:
        path: 图片文件的绝对路径（base64 编码的 UTF-8 字符串）。

    输出:
        图片文件响应；路径不存在时返回 404。
    """
    import urllib.parse
    if not path:
        raise HTTPException(status_code=404, detail="缺少 path 参数")
    try:
        real_path = base64.b64decode(path).decode("utf-8")
    except Exception:
        # 兼容未编码的原始路径
        real_path = urllib.parse.unquote(path)
    if not real_path or not os.path.isfile(real_path):
        raise HTTPException(status_code=404, detail=f"背景图片文件不存在: {real_path}")
    ext = os.path.splitext(real_path)[1].lower()
    if ext not in (".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".svg"):
        raise HTTPException(status_code=403, detail="不支持的图片格式")
    mime, _ = mimetypes.guess_type(real_path)
    return FileResponse(real_path, media_type=mime or "image/png")


@app.get("/image/{asset_name}")
def serve_image_asset(asset_name: str) -> FileResponse:
    """提供 web/image 目录下的静态图片资源。

    输入:
        asset_name: 图片文件名。

    输出:
        图片文件响应。
    """
    path = (_IMAGE_DIR / asset_name).resolve()
    try:
        path.relative_to(_IMAGE_DIR.resolve())
    except ValueError as exc:
        raise HTTPException(status_code=404, detail="图片资源不存在") from exc
    if not path.is_file():
        raise HTTPException(status_code=404, detail=f"web/image 目录下缺少 {asset_name}")
    mime, _ = mimetypes.guess_type(str(path))
    return FileResponse(path, media_type=mime or "application/octet-stream")


# ---------- 表格文本提取（XLSX/XLS → TSV，供前端按列着色的纯文本预览） ----------

@app.get("/api/table/text")
def table_text(path: str = "", sheet: str = "") -> dict[str, Any]:
    """将表格文件（XLSX / XLS）内容提取为 TSV 文本。

    输入:
        path: 表格文件绝对路径。
        sheet: 可选的工作表名（不传时取第一个工作表）。

    输出:
        {"content": str, "sheet": str, "sheets": list[str]}
        content: TSV 分隔的表格文本（供前端按列着色渲染）。
    """
    if not path or not os.path.isfile(path):
        raise HTTPException(status_code=404, detail=f"文件不存在: {path}")
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        return _xlsx_to_tsv(path, sheet)
    if ext == ".xls":
        return _xls_to_tsv(path, sheet)
    raise HTTPException(status_code=400, detail="不支持的表格格式")


def _xlsx_to_tsv(filepath: str, sheet_name: str = "") -> dict[str, Any]:
    """将 XLSX 工作表内容提取为 TSV 文本。"""
    try:
        import openpyxl
    except ImportError:
        return {"content": "(需要安装 openpyxl 才能预览 XLSX: pip install openpyxl)", "sheet": "", "sheets": []}
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    try:
        sheets = wb.sheetnames
        if not sheets:
            return {"content": "(无工作表)", "sheet": "", "sheets": []}
        selected = sheet_name if sheet_name and sheet_name in sheets else sheets[0]
        ws = wb[selected]
        lines = []
        for row in ws.iter_rows(values_only=True):
            vals = ["" if c is None else str(c) for c in row]
            lines.append("\t".join(vals))
    finally:
        wb.close()
    return {"content": "\n".join(lines), "sheet": selected, "sheets": sheets}


def _xls_to_tsv(filepath: str, sheet_name: str = "") -> dict[str, Any]:
    """将 XLS 工作表内容提取为 TSV 文本。"""
    try:
        import xlrd
    except ImportError:
        return {"content": "(需要安装 xlrd 才能预览 XLS: pip install xlrd)", "sheet": "", "sheets": []}
    wb = xlrd.open_workbook(filepath)
    sheets = wb.sheet_names()
    if not sheets:
        return {"content": "(无工作表)", "sheet": "", "sheets": []}
    selected = sheet_name if sheet_name and sheet_name in sheets else sheets[0]
    ws = wb.sheet_by_name(selected)
    lines = []
    for r in range(ws.nrows):
        vals = []
        for c in range(ws.ncols):
            v = ws.cell_value(r, c)
            if ws.cell_type(r, c) == xlrd.XL_CELL_DATE:
                try:
                    v = xlrd.xldate_as_datetime(v, wb.datemode)
                except Exception:
                    pass
            vals.append("" if v is None else str(v))
        lines.append("\t".join(vals))
    return {"content": "\n".join(lines), "sheet": selected, "sheets": sheets}


@app.get("/")
def index_page() -> FileResponse:
    """返回前端主页面 index.html。"""
    return FileResponse(_WEB_DIR / "index.html")


# ==================== 定时任务（Cron）API ====================


def _serialize_cron_messages(task_id: str) -> list[dict[str, Any]]:
    """序列化定时任务的执行记录消息，结构与 _serialize_messages 一致。

    复用 _serialize_content 处理附件 data URL，工具调用记录从统一存储的
    cron 作用域记录仓库读取（cron 任务存于 history/cron/{task_id}/，
    与主进程 tool_call_recorder.TOOL_CALLING_ROOT 指向的 sessions 目录隔离）。
    """
    from agent.cron.storage import get_record_repository
    record_repo = get_record_repository()
    msgs = record_repo.load_messages(task_id)

    out: list[dict[str, Any]] = []
    for m in msgs:
        role = m.get("role")
        content = m.get("content")
        meta = dict(m.get("meta") or {})
        meta.setdefault("output_type", "text")
        serialized_content = _serialize_content(content, meta=meta, role=role)
        out.append({"role": role, "content": serialized_content, "meta": meta})

    # 为每条带 turn_id 的 assistant 消息挂载工具调用记录（统一存储，后端无关）
    for m in out:
        if m.get("role") != "assistant":
            continue
        turn_id = (m.get("meta") or {}).get("turn_id", "")
        if not turn_id:
            continue
        rec = record_repo.get_turn_record(task_id, turn_id)
        if not rec:
            continue
        tool_calls = rec.get("tool_calls") or []
        m_meta = m.setdefault("meta", {})
        if tool_calls:
            m_meta["tool_calls"] = tool_calls
        # 附加 turn 元信息（started_at/ended_at/duration）
        turn_meta = {
            k: rec.get(k) for k in ("started_at", "ended_at", "duration")
            if rec.get(k) is not None
        }
        if turn_meta:
            m_meta["turn_meta"] = turn_meta
    return out


def _build_trigger_from_request(body: dict[str, Any]):
    """从请求体构建 Trigger，校验必填字段。"""
    from agent.cron.models import Trigger
    trig_body = body.get("trigger") or {}
    t_type = str(trig_body.get("type", "cron")).lower()
    if t_type not in ("cron", "interval", "once"):
        raise HTTPException(status_code=400, detail=f"不支持的触发类型: {t_type}")
    trig = Trigger(
        type=t_type,
        cron=str(trig_body.get("cron", "") or ""),
        interval_seconds=int(trig_body.get("interval_seconds", 0) or 0),
        run_at=str(trig_body.get("run_at", "") or ""),
    )
    # 基本校验
    if t_type == "cron" and not trig.cron:
        raise HTTPException(status_code=400, detail="cron 类型需提供 cron 表达式")
    if t_type == "interval" and trig.interval_seconds <= 0:
        raise HTTPException(status_code=400, detail="interval 类型需提供大于 0 的 interval_seconds")
    if t_type == "once" and not trig.run_at:
        raise HTTPException(status_code=400, detail="once 类型需提供 run_at 时间")
    return trig


@app.get("/api/cron/period")
def api_cron_period_get() -> dict[str, Any]:
    """获取定时任务时间段长度（分钟），供前端设置与冲突检测使用。"""
    from agent.cron.scheduler import get_cron_time_period
    return {"period_minutes": get_cron_time_period()}


@app.post("/api/cron/period")
async def api_cron_period_set(request: Request) -> dict[str, Any]:
    """设置定时任务时间段长度（分钟），热更新调度器内存值并持久化到 env_config.json。"""
    body = await request.json()
    raw = body.get("period_minutes")
    try:
        val = float(raw)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="period_minutes 必须为数字")
    if val <= 0:
        raise HTTPException(status_code=400, detail="时间段长度必须为正数")
    from agent.cron.scheduler import set_cron_time_period
    eff = set_cron_time_period(val)
    return {"period_minutes": eff}


@app.post("/api/cron/check-conflict")
async def api_cron_check_conflict(request: Request) -> dict[str, Any]:
    """新建/修改任务前检测时段冲突。

    入参:
        ``{trigger: {type, cron, interval_seconds, run_at}, exclude_task_id?: "..."}``
        trigger 字段与新建/修改接口一致；exclude_task_id 用于修改时排除自身。

    出参:
        ``{conflict: bool, conflict_with: [name,...], period_minutes: N, next_fire: "ISO"|""}``
        若候选任务下次触发时刻落入某现有任务时间点前后一个时段窗口内则 conflict=True。
    """
    body = await request.json()
    from agent.cron.scheduler import detect_time_period_conflict
    # 候选任务的下一次触发时刻（复用 trigger 校验）
    cand_trigger = _build_trigger_from_request(body)
    exclude_id = str(body.get("exclude_task_id", "") or "") or None
    return detect_time_period_conflict(cand_trigger, exclude_task_id=exclude_id)


@app.get("/api/cron")
def api_cron_list() -> dict[str, Any]:
    """列出所有定时任务（含状态摘要）。"""
    from agent.cron.storage import get_task_repository
    repo = get_task_repository()
    tasks = repo.list_tasks()
    # 附加运行中状态（调度器内存态）
    try:
        from agent.cron.scheduler import get_scheduler
        sched = get_scheduler()
    except Exception:
        sched = None
    items = []
    for t in tasks:
        d = t.summary()
        d["is_running"] = bool(sched and sched.is_running(t.task_id))
        items.append(d)
    return {"tasks": items}


@app.post("/api/cron")
async def api_cron_create(request: Request) -> dict[str, Any]:
    """新建定时任务。"""
    body = await request.json()
    from agent.cron.models import CronTask, now_iso
    from agent.cron.storage import get_task_repository, new_task_id
    from agent.cron.scheduler import _compute_next_run
    from datetime import datetime

    trig = _build_trigger_from_request(body)
    task = CronTask(
        task_id=new_task_id(),
        name=str(body.get("name", "") or "未命名任务"),
        prompt=str(body.get("prompt", "") or ""),
        trigger=trig,
        enabled=bool(body.get("enabled", True)),
        timeout_seconds=int(body.get("timeout_seconds", 300) or 300),
        created_at=now_iso(),
        last_status="pending",
    )
    task.next_run_at = _compute_next_run(trig, datetime.now())
    get_task_repository().save_task(task)
    return {"task": task.to_dict()}


@app.get("/api/cron/{task_id}")
def api_cron_detail(task_id: str) -> dict[str, Any]:
    """获取任务详情（含完整执行历史）。"""
    from agent.cron.storage import get_task_repository
    from agent.cron.scheduler import get_scheduler
    task = get_task_repository().get_task(task_id)
    if task is None:
        raise HTTPException(status_code=404, detail="任务不存在")
    d = task.to_dict()
    try:
        d["is_running"] = get_scheduler().is_running(task_id)
    except Exception:
        d["is_running"] = False
    return {"task": d}


@app.put("/api/cron/{task_id}")
async def api_cron_update(task_id: str, request: Request) -> dict[str, Any]:
    """修改任务配置。"""
    body = await request.json()
    from agent.cron.storage import get_task_repository
    from agent.cron.scheduler import _compute_next_run, get_scheduler
    from datetime import datetime

    repo = get_task_repository()
    task = repo.get_task(task_id)
    if task is None:
        raise HTTPException(status_code=404, detail="任务不存在")

    if "name" in body:
        task.name = str(body["name"] or "")
    if "prompt" in body:
        task.prompt = str(body["prompt"] or "")
    if "timeout_seconds" in body:
        task.timeout_seconds = int(body["timeout_seconds"] or 300)
    if "enabled" in body:
        task.enabled = bool(body["enabled"])
    if "trigger" in body:
        task.trigger = _build_trigger_from_request(body)

    # 重算 next_run_at
    if task.enabled:
        task.next_run_at = _compute_next_run(task.trigger, datetime.now())
    else:
        task.next_run_at = ""
    repo.save_task(task)
    return {"task": task.to_dict()}


@app.delete("/api/cron/{task_id}")
def api_cron_delete(task_id: str) -> dict[str, Any]:
    """删除任务（含目录；运行中先 kill）。"""
    from agent.cron.storage import get_task_repository
    from agent.cron.scheduler import get_scheduler
    try:
        get_scheduler().kill_task(task_id)
    except Exception:
        pass
    get_task_repository().delete_task(task_id)
    return {"deleted": task_id}


@app.post("/api/cron/{task_id}/toggle")
def api_cron_toggle(task_id: str) -> dict[str, Any]:
    """启用/禁用任务。"""
    from agent.cron.storage import get_task_repository
    from agent.cron.scheduler import _compute_next_run, get_scheduler
    from datetime import datetime

    repo = get_task_repository()
    task = repo.get_task(task_id)
    if task is None:
        raise HTTPException(status_code=404, detail="任务不存在")
    task.enabled = not task.enabled
    if task.enabled:
        task.next_run_at = _compute_next_run(task.trigger, datetime.now())
        task.last_status = "pending" if task.last_status == "disabled" else task.last_status
    else:
        task.next_run_at = ""
        task.last_status = "disabled"
    repo.save_task(task)
    return {"task": task.summary()}


@app.post("/api/cron/{task_id}/run")
async def api_cron_run(task_id: str, request: Request) -> dict[str, Any]:
    """立即运行任务 / 手动追问。

    请求体可选 {prompt: "..."}；不传则使用任务原始 prompt。
    """
    try:
        body = await request.json()
    except Exception:
        body = {}
    prompt = body.get("prompt")
    from agent.cron.scheduler import get_scheduler
    sched = get_scheduler()
    if sched.is_running(task_id):
        raise HTTPException(status_code=409, detail="任务正在运行中")
    ok = sched.spawn_now(task_id, prompt=prompt if prompt else None)
    if not ok:
        raise HTTPException(status_code=400, detail="无法启动任务（任务不存在或正在运行）")
    return {"started": True, "task_id": task_id}


@app.get("/api/cron/{task_id}/messages")
def api_cron_messages(task_id: str) -> dict[str, Any]:
    """加载任务执行记录消息（与 /api/sessions/{id}/messages 同构）。"""
    from agent.cron.storage import get_task_repository
    task = get_task_repository().get_task(task_id)
    if task is None:
        raise HTTPException(status_code=404, detail="任务不存在")
    messages = _serialize_cron_messages(task_id)
    return {
        "task_id": task_id,
        "messages": messages,
        "last_status": task.last_status,
        "last_error": task.last_error,
    }


@app.get("/api/cron/{task_id}/live/stream")
async def api_cron_live_stream(task_id: str, request: Request):
    """SSE 实时推送任务执行的工具调用流（供前端复用 toolcalls.js 渲染）。"""
    from agent.cron import live as cron_live

    q = cron_live.subscribe(task_id)

    async def event_stream():
        try:
            # 初始快照：若任务正在运行，尝试读取子进程最近回传的状态
            try:
                from agent.cron.storage import get_task_repository
                task = get_task_repository().get_task(task_id)
                running = bool(task and task.last_status == "running")
            except Exception:
                running = False
            initial = {"tool_calls": [], "reflections": [], "started_at": None, "running": running}
            yield f"event: update\ndata: {json.dumps(initial, ensure_ascii=False)}\n\n"

            while True:
                if await request.is_disconnected():
                    break
                try:
                    msg = await asyncio.to_thread(q.get, True, 3.0)
                except _queue_module.Empty:
                    yield f"event: ping\ndata: \n\n"
                    continue
                msg_type = msg.get("type", "update")
                data = msg.get("data", {})
                yield f"event: {msg_type}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n"
                if msg_type == "finished":
                    break
        finally:
            cron_live.unsubscribe(task_id, q)

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


@app.post("/api/cron/{task_id}/abort")
def api_cron_abort(task_id: str) -> dict[str, Any]:
    """终止运行中的任务子进程。"""
    from agent.cron.scheduler import get_scheduler
    ok = get_scheduler().kill_task(task_id)
    if not ok:
        raise HTTPException(status_code=400, detail="任务未在运行")
    return {"aborted": True, "task_id": task_id}


@app.post("/api/cron/internal/live")
async def api_cron_internal_live(request: Request) -> dict[str, Any]:
    """子进程回传 live 快照（仅 127.0.0.1）。"""
    body = await request.json()
    task_id = body.get("task_id", "")
    snapshot = body.get("snapshot", {})
    if not task_id:
        raise HTTPException(status_code=400, detail="缺少 task_id")
    from agent.cron import live as cron_live
    cron_live.push_snapshot(task_id, snapshot)
    return {"ok": True}


@app.post("/api/cron/internal/finished")
async def api_cron_internal_finished(request: Request) -> dict[str, Any]:
    """子进程回传最终状态（仅 127.0.0.1）。

    更新任务状态、追加执行记录、通知前端 SSE 订阅者、重算下次运行时刻。
    """
    body = await request.json()
    task_id = body.get("task_id", "")
    status = body.get("status", "completed")
    error = body.get("error", "")
    turn_id = body.get("turn_id", "")
    if not task_id:
        raise HTTPException(status_code=400, detail="缺少 task_id")

    from agent.cron.storage import get_task_repository
    from agent.cron.models import now_iso
    from agent.cron import live as cron_live
    from agent.cron.scheduler import _compute_next_run, get_scheduler
    from datetime import datetime

    repo = get_task_repository()
    task = repo.get_task(task_id)
    started_at = task.last_run_at if task else ""
    ended_at = now_iso()

    # 更新状态（last_run_at 保留为本次开始时间）
    repo.update_status(task_id, status, error=error, last_run_at=started_at)
    # 追加执行记录
    repo.append_execution(task_id, {
        "started_at": started_at,
        "ended_at": ended_at,
        "status": status,
        "error": error,
        "trigger": "manual" if (body.get("trigger") == "manual") else "schedule",
        "turn_id": turn_id,
    })

    # 重算下次运行时刻（schedule 触发的 cron/interval）
    task = repo.get_task(task_id)
    if task and task.enabled and task.trigger.type in ("cron", "interval"):
        new_nxt = _compute_next_run(task.trigger, datetime.now())
        repo.update_status(task_id, task.last_status, next_run_at=new_nxt)
    elif task and task.trigger.type == "once" and status == "completed":
        # once 类型成功后不再调度
        repo.update_status(task_id, task.last_status, next_run_at="")

    # 通知前端
    cron_live.push_finished(task_id, {"task_id": task_id, "status": status, "error": error, "turn_id": turn_id})
    return {"ok": True}


@app.get("/app.css")
def serve_css() -> FileResponse:
    """返回样式文件 app.css。"""
    return FileResponse(_WEB_DIR / "app.css")


@app.get("/app.js")
def serve_js() -> FileResponse:
    """返回前端脚本 app.js。"""
    return FileResponse(_WEB_DIR / "app.js")


@app.get("/js/{filename:path}")
def serve_js_module(filename: str) -> FileResponse:
    """返回前端 JS 模块文件。"""
    return FileResponse(_WEB_DIR / "js" / filename)


@app.get("/css/{filename:path}")
def serve_css_module(filename: str) -> FileResponse:
    """返回前端 CSS 模块文件。"""
    return FileResponse(_WEB_DIR / "css" / filename)


@app.get("/fonts/{filename}")
def serve_font(filename: str) -> FileResponse:
    """返回本地字体文件（woff2，本地化后不再依赖 Google Fonts CDN）。

    带路径包含校验，防止目录穿越。
    """
    fonts_root = (_WEB_DIR / "fonts").resolve()
    font_path = (fonts_root / filename).resolve()
    if not str(font_path).startswith(str(fonts_root) + os.sep) or not font_path.is_file():
        raise HTTPException(status_code=404, detail="字体文件不存在")
    mime, _ = mimetypes.guess_type(str(font_path))
    return FileResponse(font_path, media_type=mime or "font/woff2")


@app.get("/packages/{filename:path}")
def serve_packages(filename: str) -> FileResponse:
    """返回本地 vendor 库与 node_modules（marked/dompurify 本地化 + Monaco 离线兜底）。

    带路径包含校验，防止目录穿越。
    """
    pkg_root = (_WEB_DIR / "packages").resolve()
    target = (pkg_root / filename).resolve()
    if not str(target).startswith(str(pkg_root) + os.sep) or not target.is_file():
        raise HTTPException(status_code=404, detail="文件不存在")
    mime, _ = mimetypes.guess_type(str(target))
    return FileResponse(target, media_type=mime or "application/octet-stream")